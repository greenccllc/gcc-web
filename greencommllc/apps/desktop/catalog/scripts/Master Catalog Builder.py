"""
Master Catalog Builder v5 — Live store.ui.com price refresh + new UniFi items (E7, UCG-Industrial, ENVR Core, AI Multi Sensor 4, AI-Key, AI-Port, AI-PTZ-Precision, G6 Pro cameras, plus switch/gateway additions).

Adds:
- Full UniFi catalog: all current Gateways, Switches, APs, Protect cameras, NVRs, Access hardware
- Paxton10 controllers/readers
- Avigilon Alta (OpenPath) controllers + readers + video intercom
- Poly Trio C60 + 8800 conference phones
- Crestron DM-NVX encoder/decoder/wallplate
- Clear-Com MS-702, HelixNet HMS-4X, beltpacks
- SonicWave 231c, 641, 681 access points

Prices verified via web search (see chat history for sources):
UniFi MSRPs from ui.com/ui-care and store.ui.com; Avigilon Alta from TIP MSRP PDF;
Paxton from paxton-access.com price change notice; SonicWave from SonicGuard/Walmart;
Poly Trio from HP site / retail; Crestron DM-NVX from Creation Networks;
Clear-Com from Markertek; SonicWall NSa 2700 from SonicGuard.

Residential catalog unchanged - preserved from v3 exactly.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BRAND_GREEN = "2E7D32"
LABOR_COST_PER_HR = 63.26
LABOR_SALE_PER_HR = 100.00

OUT_DIR = "/home/user/AI_DRIVE_OUTPUT"
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================================
# COMMERCIAL EQUIPMENT (Sheet 1)
# Format: (name, units, cost, sale_price, model, description, purchase_link)
# ============================================================================
commercial_equipment = [
    # ========== UNIFI GATEWAYS (full line) ==========
    ("UniFi Cloud Gateway Ultra", "ea", 129, 179, "UCG-Ultra", "Compact 1G cloud gateway, 50+ devices", "https://store.ui.com/us/en/products/ucg-ultra"),
    ("UniFi Cloud Gateway Max", "ea", 279, 386, "UCG-Max", "2.5G cloud gateway, 30+ devices / 300+ clients", "https://store.ui.com/us/en/products/ucg-max"),
    ("UniFi Cloud Gateway Max (512GB)", "ea", 199, 277, "UCG-Max-NS", "2.5G cloud gateway w/ 512GB NVMe for Protect", "https://store.ui.com/us/en/products/ucg-max-ns"),
    ("UniFi Cloud Gateway Fiber", "ea", 279, 389, "UCG-Fiber", "10G desktop gateway w/ integrated 4-port 2.5GbE PoE", "https://store.ui.com/us/en/products/ucg-fiber"),
    ("UniFi Dream Router 7", "ea", 279, 389, "UDR7", "WiFi 7 all-in-one gateway + AP + switch", "https://store.ui.com/us/en/products/udr7"),
    ("UniFi Dream Machine SE", "ea", 499, 695, "UDM-SE", "1U rack-mount 1G cloud gateway w/ Protect", "https://store.ui.com/us/en/products/udm-se"),
    ("UniFi Dream Machine Pro", "ea", 379, 525, "UDM-Pro", "1U rack-mount 1G cloud gateway + NVR", "https://store.ui.com/us/en/products/udm-pro"),
    ("UniFi Dream Machine Pro Max", "ea", 599, 825, "UDM-Pro-Max", "1U 10G cloud gateway w/ 200+ device support", "https://store.ui.com/us/en/products/udm-pro-max"),
    ("UniFi Enterprise Fortress Gateway", "ea", 1999, 2695, "EFG", "2U 25G enterprise firewall/gateway, 1000+ clients", "https://store.ui.com/us/en/products/efg"),

    # ========== UNIFI SWITCHES — Standard tier ==========
    ("UniFi Switch Flex Mini", "ea", 29, 45, "USW-Flex-Mini", "5-port 1G desktop switch, PoE-powered", "https://store.ui.com/us/en/products/usw-flex-mini"),
    ("UniFi Switch Flex", "ea", 99, 145, "USW-Flex", "5-port 1G outdoor switch, PoE passthrough", "https://store.ui.com/us/en/products/usw-flex"),
    ("UniFi Switch Lite 8 PoE", "ea", 109, 155, "USW-Lite-8-PoE", "8-port 1G L2 switch, 4 PoE+ (52W)", "https://store.ui.com/us/en/products/usw-lite-8-poe"),
    ("UniFi Switch Lite 16 PoE", "ea", 199, 279, "USW-Lite-16-PoE", "16-port 1G L2 switch, 8 PoE+ (45W)", "https://store.ui.com/us/en/products/usw-lite-16-poe"),
    ("UniFi Switch 24", "ea", 225, 315, "USW-24", "24-port 1G L2 switch, 2 SFP, no PoE", "https://store.ui.com/us/en/products/usw-24"),
    ("UniFi Switch 16 PoE", "ea", 299, 419, "USW-16-PoE", "16-port 1G L2 switch, 8 PoE+ (42W)", "https://store.ui.com/us/en/products/usw-16-poe"),
    ("UniFi Switch 24 PoE", "ea", 379, 525, "USW-24-PoE", "24-port 1G L2 switch, 16 PoE (95W)", "https://store.ui.com/us/en/products/usw-24-poe"),
    ("UniFi Switch 48 PoE", "ea", 589, 825, "USW-48-PoE", "48-port 1G L2 switch, 32 PoE (195W)", "https://store.ui.com/us/en/products/usw-48-poe"),

    # ========== UNIFI SWITCHES — Pro tier (L3) ==========
    ("UniFi Switch Pro 24", "ea", 399, 555, "USW-Pro-24", "24-port 1G L3 switch, 2 10G SFP+", "https://store.ui.com/us/en/products/usw-pro-24"),
    ("UniFi Switch Pro 24 PoE", "ea", 699, 975, "USW-Pro-24-PoE", "24-port 1G L3 switch, 16 PoE+ / 8 PoE++ (400W)", "https://store.ui.com/us/en/products/usw-pro-24-poe"),
    ("UniFi Switch Pro 48 PoE", "ea", 1099, 1535, "USW-Pro-48-PoE", "48-port 1G L3 switch, 40 PoE+ / 8 PoE++ (600W)", "https://store.ui.com/us/en/products/usw-pro-48-poe"),
    ("UniFi Switch Pro Max 16 PoE", "ea", 399, 555, "USW-Pro-Max-16-PoE", "16-port 2.5G Etherlighting L3, 8 PoE++ (240W)", "https://store.ui.com/us/en/products/usw-pro-max-16-poe"),
    ("UniFi Switch Pro Max 24 PoE", "ea", 799, 1100, "USW-Pro-Max-24-PoE", "24-port 2.5G Etherlighting L3, PoE++ (400W)", "https://store.ui.com/us/en/products/usw-pro-max-24-poe"),
    ("UniFi Switch Pro Max 48 PoE", "ea", 1299, 1804, "USW-Pro-Max-48-PoE", "48-port 2.5G Etherlighting L3, PoE++ (720W)", "https://store.ui.com/us/en/products/usw-pro-max-48-poe"),
    ("UniFi Switch Pro XG 10 PoE", "ea", 699, 976, "USW-Pro-XG-10-PoE", "8-port 10G RJ45 + 2 25G SFP28 PoE++ (240W)", "https://store.ui.com/us/en/products/usw-pro-xg-10-poe"),
    ("UniFi Switch Pro XG 24", "ea", 1099, 1536, "USW-Pro-XG-24", "24-port 10G RJ45 + 2 25G SFP28 (no PoE)", "https://store.ui.com/us/en/products/usw-pro-xg-24"),
    ("UniFi Switch Pro Aggregation", "ea", 899, 1251, "USW-Pro-Aggregation", "28-port 10G/25G SFP+/SFP28 aggregation", "https://store.ui.com/us/en/products/usw-pro-aggregation"),

    # ========== UNIFI SWITCHES — Enterprise tier ==========
    ("UniFi Switch Enterprise 8 PoE", "ea", 479, 667, "USW-Enterprise-8-PoE", "8-port 2.5G Enterprise L3 PoE+++ (120W)", "https://store.ui.com/us/en/products/usw-enterprise-8-poe"),
    ("UniFi Switch Enterprise 24 PoE", "ea", 799, 1115, "USW-Enterprise-24-PoE", "24-port 2.5G Enterprise L3 PoE+++ (400W)", "https://store.ui.com/us/en/products/usw-enterprise-24-poe"),
    ("UniFi Switch Enterprise 48 PoE", "ea", 1599, 2225, "USW-Enterprise-48-PoE", "48-port 2.5G Enterprise L3 PoE+++ (720W)", "https://store.ui.com/us/en/products/usw-enterprise-48-poe"),
    ("UniFi Switch Enterprise Campus 48 PoE", "ea", 2999, 4175, "USW-EnterpriseCampus-48-PoE", "48-port 10G RJ45 PoE+++ w/ 25G SFP28 uplinks", "https://store.ui.com/us/en/products/usw-enterprisecampus-48-poe"),

    # ========== UNIFI ACCESS POINTS — WiFi 7 ==========
    ("UniFi AP U7 In-Wall", "ea", 149, 209, "U7-IW", "Wall-mounted WiFi 7 AP, 4-port switch integrated", "https://store.ui.com/us/en/products/u7-iw"),
    ("UniFi AP U7 Pro Wall", "ea", 199, 279, "U7-Pro-Wall", "Wall-mounted 6GHz-ready WiFi 7 AP, 6 streams", "https://store.ui.com/us/en/products/u7-pro-wall"),
    ("UniFi AP U7 Pro", "ea", 189, 265, "U7-Pro", "Ceiling-mount WiFi 7 AP, 6 streams, 6GHz", "https://store.ui.com/us/en/products/u7-pro"),
    ("UniFi AP E7 (Enterprise WiFi 7)", "ea", 499, 699, "E7", "Enterprise WiFi 7 AP, 10 streams, 10 GbE uplink + GbE failover", "https://store.ui.com/us/en/products/e7"),
    ("UniFi AP E7 Campus (WiFi 7 directional)", "ea", 799, 1119, "E7-Campus", "Enterprise WiFi 7, 10-stream, PRISM RF filtering, directional", "https://store.ui.com/us/en/collections/pro-e7-campus"),
    ("UniFi AP E7 Audience (WiFi 7 high-density)", "ea", 1999, 2795, "E7-Audience", "Enterprise 12-stream 5/6 GHz WiFi 7 for high-density venues", "https://store.ui.com/us/en/collections/pro-e7-audience"),
    ("UniFi AP U7 Pro XGS (WiFi 7 8-stream)", "ea", 299, 419, "U7-Pro-XGS", "Ceiling WiFi 7 AP, 8 streams, spectral scan, 10/5/2.5/1 GbE", "https://store.ui.com/us/en/products/u7-pro-xgs"),
    ("UniFi AP U7 Pro XG (WiFi 7 6-stream)", "ea", 199, 279, "U7-Pro-XG", "Ceiling WiFi 7 AP, 6 streams, 10/5/2.5/1 GbE", "https://store.ui.com/us/en/products/u7-pro-xg"),
    ("UniFi AP U7 LR (Long-Range WiFi 7)", "ea", 159, 225, "U7-LR", "Ceiling WiFi 7 AP, 5 streams, extended range", "https://store.ui.com/us/en/products/u7-lr"),
    ("UniFi AP U7 Lite (WiFi 7 compact)", "ea", 99, 139, "U7-Lite", "Compact ceiling WiFi 7 AP, 4 streams, 2.5 GbE", "https://store.ui.com/us/en/products/u7-lite"),
    ("UniFi AP U7 Mesh (indoor/outdoor)", "ea", 199, 279, "U7-Mesh", "Indoor/outdoor WiFi 7 AP w/ long-range antenna", "https://store.ui.com/us/en/products/u7-mesh"),
    ("UniFi AP U7 Pro Max", "ea", 279, 389, "U7-Pro-Max", "Ceiling-mount WiFi 7 AP, 8 streams, 6GHz", "https://store.ui.com/us/en/products/u7-pro-max"),
    ("UniFi AP U7 Pro XG Wall", "ea", 279, 391, "U7-Pro-XG-Wall", "Wall-mount WiFi 7 AP, 10GbE uplink, 6GHz", "https://store.ui.com/us/en/products/u7-pro-xg-wall"),
    ("UniFi AP U7 Outdoor", "ea", 199, 279, "U7-Outdoor", "Outdoor WiFi 7 AP, IP67 rated", "https://store.ui.com/us/en/products/u7-outdoor"),
    ("UniFi AP U7 Pro Outdoor", "ea", 179, 250, "U7-Pro-Outdoor", "Outdoor IP67 WiFi 7 AP w/ extended range antennas", "https://store.ui.com/us/en/products/u7-pro-outdoor"),

    # ========== UNIFI ACCESS POINTS — WiFi 6E ==========
    ("UniFi AP U6 Enterprise", "ea", 279, 391, "U6-Enterprise", "Ceiling-mount WiFi 6E AP, 6GHz, 10GbE uplink", "https://store.ui.com/us/en/products/u6-enterprise"),
    ("UniFi AP U6 Enterprise In-Wall", "ea", 299, 419, "U6-Enterprise-IW", "Wall-mount WiFi 6E AP w/ 4-port switch", "https://store.ui.com/us/en/products/u6-enterprise-iw"),

    # ========== UNIFI ACCESS POINTS — WiFi 6 ==========
    ("UniFi AP U6+", "ea", 129, 179, "U6-Plus", "Ceiling-mount WiFi 6 AP, 2x2 MIMO", "https://store.ui.com/us/en/products/u6-plus"),
    ("UniFi AP U6 Pro", "ea", 159, 223, "U6-Pro", "Ceiling-mount WiFi 6 AP, 4x4 MU-MIMO", "https://store.ui.com/us/en/products/u6-pro"),
    ("UniFi AP U6 Long-Range", "ea", 179, 249, "U6-LR", "Ceiling-mount WiFi 6 AP, long-range antenna", "https://store.ui.com/us/en/products/u6-lr"),
    ("UniFi AP U6 Mesh", "ea", 179, 251, "U6-Mesh", "Indoor/outdoor WiFi 6 AP, 6-stream mesh", "https://store.ui.com/us/en/products/u6-mesh"),
    ("UniFi AP U6 Mesh Pro", "ea", 199, 279, "U6-Mesh-Pro", "Outdoor WiFi 6 AP, integrated super antenna", "https://store.ui.com/us/en/products/u6-mesh-pro"),
    ("UniFi AP U6 In-Wall", "ea", 179, 251, "U6-IW", "Wall-mount WiFi 6 AP, 4-port switch integrated", "https://store.ui.com/us/en/products/u6-iw"),
    ("UniFi AP U6 Extender", "ea", 149, 209, "U6-Extender", "Plug-in WiFi 6 mesh extender", "https://store.ui.com/us/en/products/u6-extender"),

    # ========== UNIFI PROTECT CAMERAS — G6 (current flagship, 2025) ==========
    ("UniFi Protect G6 Bullet", "ea", 199, 279, "UVC-G6-Bullet", "4K bullet cam, Multi-TOPS AI, PoE", "https://store.ui.com/us/en/products/uvc-g6-bullet"),
    ("UniFi Protect G6 Turret", "ea", 199, 279, "UVC-G6-Turret", "4K turret cam, tamper-resistant, PoE", "https://store.ui.com/us/en/products/uvc-g6-turret"),
    ("UniFi Protect G6 Instant", "ea", 179, 249, "UVC-G6-Instant", "Compact 2K cam w/ SD card, PoE or USB-C", "https://store.ui.com/us/en/products/uvc-g6-instant"),
    ("UniFi Protect G6 Pro Bullet", "ea", 479, 669, "UVC-G6-Pro-Bullet", "4K PoE+ bullet, Multi-TOPS AI, 2.36x zoom, 1/1.2\" CMOS, 40m IR", "https://store.ui.com/us/en/products/uvc-g6-pro-bullet"),
    ("UniFi Protect G6 Pro Turret", "ea", 479, 669, "UVC-G6-Pro-Turret", "4K PoE+ turret, Multi-TOPS AI, 2.36x zoom, 1/1.2\" CMOS, 40m IR", "https://store.ui.com/us/en/products/uvc-g6-pro-turret"),
    ("UniFi Protect G6 Pro Dome", "ea", 499, 699, "UVC-G6-Pro-Dome", "4K PoE+ vandal dome, Multi-TOPS AI, 2.36x zoom, IK10", "https://store.ui.com/us/en/products/uvc-g6-pro-dome"),
    ("UniFi Protect G6 Pro 360 (12MP panoramic)", "ea", 499, 699, "UVC-G6-Pro-360", "12MP 360° vandal panoramic, digital PTZ, IK10", "https://store.ui.com/us/en/products/uvc-g6-pro-360"),
    ("UniFi Protect G6 180 (16MP dual-sensor)", "ea", 299, 419, "UVC-G6-180", "16MP 180° dual-sensor PoE+, AI, 20 FPS", "https://store.ui.com/us/en/products/uvc-g6-180"),
    ("UniFi Protect G6 PTZ", "ea", 399, 559, "UVC-G6-PTZ", "4K dual-lens PTZ, 10x hybrid zoom, AI, 2-way audio", "https://store.ui.com/us/en/products/uvc-g6-ptz"),
    ("UniFi Protect AI Multi Sensor 4 (32MP)", "ea", 1799, 2519, "UVC-AI-MS-4", "32MP PoE++ multi-sensor, 2.33x zoom, 360° IR, 4 adjustable lenses", "https://store.ui.com/us/en/products/uvc-ai-ms-4"),
    ("UniFi Protect AI LPR (license plate)", "ea", 499, 699, "UVC-AI-LPR", "4K LPR cam, 3x zoom, up to 90 km/h plate capture", "https://store.ui.com/us/en/products/uvc-ai-lpr"),
    ("UniFi Protect AI PTZ Industrial (22x zoom)", "ea", 1299, 1815, "UVC-AI-PTZ", "Industrial 4K PoE++ PTZ, 22x zoom, AI, 100m IR", "https://store.ui.com/us/en/products/uvc-ai-ptz"),
    ("UniFi Protect AI PTZ Precision (31x LiDAR)", "ea", 1999, 2795, "UVC-AI-PTZ-Precision", "Industrial 4K PTZ, 31x zoom, LiDAR autofocus, 100m IR", "https://store.ui.com/us/en/products/uvc-ai-ptz-precision"),

    # ========== UNIFI PROTECT CAMERAS — G5 (available legacy) ==========
    ("UniFi Protect G5 Bullet", "ea", 129, 181, "UVC-G5-Bullet", "2K bullet cam, PoE, indoor/outdoor", "https://store.ui.com/us/en/products/uvc-g5-bullet"),
    ("UniFi Protect G5 Dome Ultra", "ea", 129, 179, "UVC-G5-Dome-Ultra", "Ultra-compact 2K dome, vandal-resistant", "https://store.ui.com/us/en/products/uvc-g5-dome-ultra"),
    ("UniFi Protect G5 Turret Ultra", "ea", 129, 179, "UVC-G5-Turret-Ultra", "Ultra-compact 2K turret, long-range IR", "https://store.ui.com/us/en/products/uvc-g5-turret-ultra"),
    ("UniFi Protect G5 Flex", "ea", 129, 181, "UVC-G5-Flex", "Flexible-mount 2K cam, ceiling/wall/desk", "https://store.ui.com/us/en/products/uvc-g5-flex"),
    ("UniFi Protect G5 Pro", "ea", 449, 625, "UVC-G5-Pro", "4K Pro cam, long-range, 1/1.8\" CMOS", "https://store.ui.com/us/en/products/uvc-g5-pro"),
    ("UniFi Protect G5 PTZ", "ea", 1099, 1535, "UVC-G5-PTZ", "4K PTZ, 22x optical zoom, auto-tracking", "https://store.ui.com/us/en/products/uvc-g5-ptz"),

    # ========== UNIFI PROTECT CAMERAS — AI & Panoramic ==========
    ("UniFi Protect AI Pro", "ea", 499, 696, "UVC-AI-Pro", "4K AI bullet cam, on-edge object detection", "https://store.ui.com/us/en/products/uvc-ai-pro"),
    ("UniFi Protect AI Theta", "ea", 1399, 1955, "UVC-AI-Theta", "Modular panoramic AI cam, 360° coverage", "https://store.ui.com/us/en/products/uvc-ai-theta"),
    ("UniFi Protect AI 360", "ea", 399, 559, "UVC-AI-360", "Fisheye panoramic 360° AI cam", "https://store.ui.com/us/en/products/uvc-ai-360"),

    # ========== UNIFI PROTECT — Doorbells & Entry ==========
    ("UniFi Protect G4 Doorbell Pro", "ea", 299, 419, "UVC-G4-DoorBell-Pro", "WiFi + PoE doorbell cam w/ secondary screen", "https://store.ui.com/us/en/products/uvc-g4-doorbell-pro"),
    ("UniFi Protect G4 Doorbell Pro PoE", "ea", 299, 417, "UVC-G4-DoorBell-Pro-PoE", "PoE doorbell cam w/ fingerprint + NFC", "https://store.ui.com/us/en/products/uvc-g4-doorbell-pro-poe"),
    ("UniFi Protect G6 Entry", "ea", 249, 349, "UVC-G6-Entry", "Access/Protect door entry system, 2-way audio", "https://store.ui.com/us/en/products/uvc-g6-entry"),
    ("UniFi Protect G6 Pro Entry", "ea", 499, 695, "UVC-G6-Pro-Entry", "Premium entry w/ 12MP cam + 3\" touch display", "https://store.ui.com/us/en/products/uvc-g6-pro-entry"),

    # ========== UNIFI PROTECT — NVRs ==========
    ("UniFi Network Video Recorder (4-bay)", "ea", 299, 419, "UNVR", "4-bay NVR, up to (18) 4K or (60) Full HD, 30-day storage", "https://store.ui.com/us/en/products/unvr"),
    ("UniFi Network Video Recorder Pro (7-bay)", "ea", 499, 699, "UNVR-Pro", "2U 7-bay NVR, up to (24) 4K or (70) Full HD, 60-day storage", "https://store.ui.com/us/en/products/unvr-pro"),
    ("UniFi Network Video Recorder Instant Kit", "ea", 699, 975, "UNVR-Instant", "All-in-one NVR + 4 G5 Turret Ultra + 1TB", "https://store.ui.com/us/en/products/unvr-instant"),
    ("UniFi Enterprise NVR (16-bay, 3U)", "ea", 1999, 2795, "ENVR", "3U 16-bay NVR, up to (70) 4K or (210) Full HD, RAID, hot-swap PSUs", "https://store.ui.com/us/en/products/envr"),
    ("UniFi Enterprise NVR Core (16-bay mega-scale)", "ea", 4999, 6995, "ENVR-Core", "3U mega-scale NVR, (300) 4K or (500) Full HD, SFP28 + SFF-8644", "https://store.ui.com/us/en/products/envr-core"),
    ("UniFi AI Key (edge AI appliance)", "ea", 799, 1119, "AI-Key", "Proactive AI threat detection, 1,800 smart events/hr", "https://store.ui.com/us/en/products/ai-key"),
    ("UniFi AI Port (3rd-party camera enhancer)", "ea", 199, 279, "UP-AI-Port", "AI detection/classification for any UniFi or 3rd-party cam", "https://store.ui.com/us/en/products/up-ai-port"),
    ("UniFi Protect Viewport", "ea", 199, 279, "UFP-Viewport", "HDMI hub for live multi-camera display", "https://store.ui.com/us/en/products/ufp-viewport"),
    ("UniFi CloudKey+ G2 SSD", "ea", 249, 349, "UCK-G2-SSD", "Compact UniFi console w/ pre-installed SSD, PoE", "https://store.ui.com/us/en/products/uck-g2-ssd"),

    # ========== UNIFI ACCESS (physical access control) ==========
    ("UniFi Access Hub Enterprise", "ea", 279, 388, "UA-Hub", "4-door enterprise access control hub", "https://store.ui.com/us/en/products/ua-hub"),
    ("UniFi Access Hub Door", "ea", 199, 277, "UA-Hub-Door", "Single-door access control hub", "https://store.ui.com/us/en/products/ua-hub-door"),
    ("UniFi Access Hub Door Mini", "ea", 129, 183, "UA-Hub-Door-Mini", "Compact single-door access hub", "https://store.ui.com/us/en/products/ua-hub-door-mini"),
    ("UniFi Access Gate Hub", "ea", 279, 388, "UA-Hub-Gate", "Multi-reader gate access control hub", "https://store.ui.com/us/en/products/ua-hub-gate"),
    ("UniFi Access Elevator Hub Starter Kit", "ea", 999, 1393, "UA-SK-Elevator", "Elevator floor access hub + G2 Reader", "https://store.ui.com/us/en/products/ua-sk-elevator"),
    ("UniFi Access Retrofit Hub", "ea", 229, 321, "UA-Retrofit-Hub-2", "Retrofit existing wiring to UniFi Access", "https://store.ui.com/us/en/products/ua-retrofit-hub-2"),
    ("UniFi Access Reader G3", "ea", 159, 223, "UA-G3", "Third-gen NFC/Bluetooth reader, indoor/outdoor", "https://store.ui.com/us/en/products/ua-g3"),
    ("UniFi Access Reader G3 Pro", "ea", 379, 531, "UA-G3-Pro", "Third-gen Pro reader w/ PIN + camera", "https://store.ui.com/us/en/products/ua-g3-pro"),
    ("UniFi Access Reader G3 Flex", "ea", 199, 281, "UA-G3-Flex", "Third-gen compact reader, mullion form", "https://store.ui.com/us/en/products/ua-g3-flex"),
    ("UniFi Access Reader G2 Pro", "ea", 359, 503, "UA-G2-Pro", "Second-gen Pro reader w/ touch display", "https://store.ui.com/us/en/products/ua-g2-pro"),
    ("UniFi Access Reader Lite", "ea", 69, 95, "UA-Lite", "Entry-level NFC/BT reader", "https://store.ui.com/us/en/products/ua-lite"),
    ("UniFi Access Keypad", "ea", 199, 279, "UA-Keypad", "PIN-code keypad w/ NFC support", "https://store.ui.com/us/en/products/ua-keypad"),
    ("UniFi Access Ultra", "ea", 129, 181, "UA-Ultra", "Integrated hub + reader, single-door", "https://store.ui.com/us/en/products/ua-ultra"),
    ("UniFi Access Intercom", "ea", 399, 555, "UA-Intercom", "Outdoor video intercom w/ 2-way audio", "https://store.ui.com/us/en/products/ua-intercom"),
    ("UniFi Access Intercom Viewer", "ea", 199, 279, "UA-Intercom-Viewer", "Indoor answering station for Intercom", "https://store.ui.com/us/en/products/ua-intercom-viewer"),
    ("UniFi Access Card (10-pack)", "pack", 29, 42, "UA-Card-10", "MIFARE NFC cards, 13.56MHz, 10-pack", "https://store.ui.com/us/en/products/ua-card-10"),

    # ========== CISCO MERAKI (when spec'd) ==========
    ("Cisco Meraki MX75 Security Appliance", "ea", 2150, 3025, "MX75-HW", "Cloud-managed SD-WAN firewall, 2 WAN 1G", "https://meraki.cisco.com/products/security-sd-wan/mx75"),
    ("Cisco Meraki MS225-48FP Switch", "ea", 1895, 2655, "MS225-48FP-HW", "48-port cloud-managed L2 switch, 740W PoE", "https://meraki.cisco.com/products/switches/ms225"),

    # ========== CISCO CATALYST (when spec'd) ==========
    ("Cisco Catalyst 9200L 24-Port PoE+", "ea", 2495, 3495, "C9200L-24P-4G-E", "24-port L3 PoE+ switch, 4x 1G uplinks", "https://www.cisco.com/c/en/us/products/switches/catalyst-9200-series-switches/"),

    # ========== HP ARUBA (when spec'd) ==========
    ("HP Aruba 6200F 48G PoE+", "ea", 2295, 3225, "JL727A", "48-port L3 PoE+ switch, 740W, 4 SFP+", "https://www.arubanetworks.com/products/switches/6200-series/"),

    # ========== FORTINET FORTIGATE ==========
    ("FortiGate 60F Firewall Appliance", "ea", 695, 975, "FG-60F", "NGFW hardware only, 10Gbps FW, 1.4Gbps IPS", "https://www.fortinet.com/products/next-generation-firewall"),
    ("FortiGate 100F Firewall Appliance", "ea", 2995, 4195, "FG-100F", "NGFW hardware only, 20Gbps FW, dual PSU, 22x1G RJ45", "https://www.fortinet.com/products/next-generation-firewall"),

    # ========== SONICWALL FIREWALLS ==========
    ("SonicWall TZ370 Firewall (hardware only)", "ea", 995, 1395, "02-SSC-2825", "SMB firewall appliance, 3Gbps threat prevention", "https://www.sonicguard.com/TZ370.asp"),
    ("SonicWall TZ470 Firewall (hardware only)", "ea", 1295, 1815, "02-SSC-6816", "SMB firewall appliance, 5Gbps threat prevention", "https://www.sonicguard.com/TZ470.asp"),
    ("SonicWall TZ570 Firewall (hardware only)", "ea", 1695, 2375, "02-SSC-5660", "Mid-market firewall, 8Gbps threat prevention", "https://www.sonicguard.com/TZ570.asp"),
    ("SonicWall TZ670 Firewall (hardware only)", "ea", 1595, 2235, "02-SSC-5668", "Mid-market firewall, 10Gbps threat prevention", "https://www.sonicguard.com/TZ670.asp"),
    ("SonicWall NSa 2700 Firewall (hardware only)", "ea", 2195, 3075, "02-SSC-7367", "Enterprise firewall, 11Gbps threat prevention", "https://www.sonicguard.com/NSA-2700.asp"),
    ("SonicWall AGSS License 1YR (TZ470)", "ea", 895, 1255, "02-SSC-6826", "Advanced Gateway Security Suite, 1 year", "https://www.sonicguard.com/TZ470.asp"),
    ("SonicWall AGSS License 1YR (TZ570)", "ea", 1095, 1535, "02-SSC-5666", "Advanced Gateway Security Suite, 1 year", "https://www.sonicguard.com/TZ570.asp"),

    # ========== SONICWALL SONICWAVE (WiFi when spec'd) ==========
    ("SonicWall SonicWave 231c Ceiling AP", "ea", 395, 555, "02-SSC-2434", "WiFi 5 Wave 2 ceiling AP, plenum-rated", "https://www.sonicguard.com/SonicWave.asp"),
    ("SonicWall SonicWave 641 Indoor AP", "ea", 566, 795, "03-SSC-0350", "WiFi 6 indoor AP w/ integrated security", "https://www.sonicguard.com/SonicWave.asp"),
    ("SonicWall SonicWave 681 Flagship AP", "ea", 895, 1255, "03-SSC-0451", "WiFi 6E flagship AP, tri-radio", "https://www.sonicguard.com/SonicWave.asp"),

    # ========== ACCESS CONTROL PANELS & CONTROLLERS (non-UniFi) ==========
    ("Mercury LP1502 2-Reader Controller", "ea", 425, 595, "LP1502", "Intelligent 2-reader controller, IP/Ethernet", "https://www.hidglobal.com/products/mercury-lp1502"),
    ("Mercury LP4502 16-Reader Controller", "ea", 1295, 1815, "LP4502", "Intelligent 16-reader controller, IP/Ethernet", "https://www.hidglobal.com/products/mercury-lp4502"),
    ("Axis A1601 Network Door Controller", "ea", 560, 785, "A1601", "Open-protocol 1-door controller, PoE", "https://www.axis.com/products/axis-a1601-network-door-controller"),
    ("Axis A1610 Network Door Controller", "ea", 995, 1395, "A1610", "Enterprise 2-door controller, PoE", "https://www.axis.com/products/axis-a1610-network-door-controller"),
    ("Avigilon Alta (OpenPath) Core Series 4-Door Hub", "ea", 1495, 2095, "4ENT-SYS-24V", "Cloud-managed 4-door access control hub", "https://www.avigilon.com/access-control/12-24v-4-door-core-series-smart-hub"),
    ("Avigilon Alta (OpenPath) Core Series 8-Door Hub", "ea", 2495, 3495, "8ENT-SYS-DVE2", "Cloud-managed 8-door access control hub", "https://www.avigilon.com/access-control/12-24v-8-door-core-series-smart-hub"),
    ("Avigilon Alta (OpenPath) 4-Port Expansion Board", "ea", 495, 695, "OP-EX-4E", "4-reader expansion board for Core hub", "https://www.avigilon.com/access-control/4-port-board-op-4ex4"),
    ("Avigilon Alta (OpenPath) 8-Port Expansion Board", "ea", 795, 1115, "OP-EX-8E", "8-reader expansion board for Core hub", "https://www.avigilon.com/access-control/8-port-board-op-8ex8"),
    ("Brivo ACS6100 4-Door Panel", "ea", 1895, 2655, "ACS6100", "Cloud-managed 4-door IP controller", "https://www.brivo.com/products/access-control/"),
    ("HID Aero X1100 Door Controller", "ea", 795, 1115, "X1100", "Intelligent 1-door controller, open platform", "https://www.hidglobal.com/products/aero-x1100"),
    ("Paxton10 Single-Door Controller (PoE, Metal Housing)", "ea", 495, 695, "010-495-US", "Cloud-managed PoE door controller w/ metal enclosure", "https://www.paxton-access.com/product-types/paxton10/"),
    ("Paxton10 Single-Door Controller (12/24V PSU)", "ea", 425, 595, "010-522-US", "Cloud-managed door controller w/ PSU, for standalone doors", "https://www.paxton-access.com/product-types/paxton10/"),

    # ========== CARD READERS (non-UniFi) ==========
    ("HID Signo 20 Reader (Mullion)", "ea", 225, 315, "SIGNO-20NKS", "Multi-tech mullion reader, BLE/NFC/125kHz/13.56MHz", "https://www.hidglobal.com/products/signo-reader-20"),
    ("HID Signo 20 Reader (Keypad)", "ea", 295, 415, "SIGNO-20TKS", "Multi-tech mullion reader w/ keypad", "https://www.hidglobal.com/products/signo-reader-20"),
    ("HID Signo 40 Reader (Wall)", "ea", 275, 385, "SIGNO-40NKS", "Multi-tech wall-mount reader, BLE/NFC", "https://www.hidglobal.com/products/signo-reader-40"),
    ("HID Signo 40 Reader (Keypad)", "ea", 345, 485, "SIGNO-40TKS", "Multi-tech wall-mount reader w/ keypad", "https://www.hidglobal.com/products/signo-reader-40"),
    ("HID iCLASS SE R10 Reader (Mullion)", "ea", 195, 275, "900NMNNEK00000", "iCLASS SE mullion reader, 13.56MHz", "https://www.hidglobal.com/products/iclass-se-readers"),
    ("HID iCLASS SE R40 Reader (Wall)", "ea", 265, 375, "920NMNNEK00000", "iCLASS SE wall reader, 13.56MHz", "https://www.hidglobal.com/products/iclass-se-readers"),
    ("HID MultiClass SE RP40 Reader", "ea", 295, 415, "921NTNTEK00000", "Multi-tech RP40 wall reader, Prox + iCLASS", "https://www.hidglobal.com/products/multiclass-se-readers"),
    ("Axis TA1601 I/O Relay Module", "ea", 295, 415, "TA1601", "Add-on I/O relay for Axis A1601/A1610", "https://www.axis.com/products/axis-ta1601"),
    ("Avigilon Alta Standard Smart Reader v2", "ea", 295, 415, "OP-R2X-STND", "BLE/NFC/Wiegand reader, wall form factor", "https://www.avigilon.com/access-control/"),
    ("Avigilon Alta Mullion Smart Reader v2", "ea", 295, 415, "OP-R2X-MULL", "BLE/NFC/Wiegand reader, mullion form factor", "https://www.avigilon.com/access-control/"),
    ("Avigilon Alta Standard Keypad Reader", "ea", 345, 485, "OP-RKP-STND", "BLE/NFC/Wiegand keypad reader", "https://www.avigilon.com/access-control/"),
    ("Avigilon Alta Video Intercom Reader", "ea", 795, 1115, "OP-VID-PRO-RDR", "Integrated video reader + intercom", "https://www.avigilon.com/access-control/"),
    ("Avigilon Alta Video Pro Intercom", "ea", 895, 1255, "OP-VID-PRO-INT", "Video intercom station w/ cam + audio", "https://www.avigilon.com/access-control/"),
    ("Paxton10 Bluetooth Reader + Keypad", "ea", 325, 455, "010-721-US", "Bluetooth hands-free reader w/ keypad", "https://www.paxton-access.com/product-types/paxton10/"),

    # ========== ELECTRIC LOCKS & STRIKES ==========
    ("HES 1006 Electric Strike (12/24V)", "ea", 325, 455, "1006", "Universal electric strike for cylindrical locks", "https://www.assaabloyesh.com/hes/1006/"),
    ("HES 5000C Electric Strike", "ea", 425, 595, "5000C", "Heavy-duty mortise strike, 12/24V", "https://www.assaabloyesh.com/hes/5000c/"),
    ("HES 9600 Rim Exit Strike", "ea", 625, 875, "9600", "Fire-rated strike for rim exit devices", "https://www.assaabloyesh.com/hes/9600/"),
    ("Von Duprin 6111 Fire-Rated Strike", "ea", 695, 975, "6111", "UL fire-rated electric strike, 24VDC", "https://www.allegion.com/corp/en/brands/von-duprin.html"),
    ("Securitron M62 Magnetic Lock (600lb)", "ea", 295, 415, "M62", "Surface-mount maglock, 600lb holding force", "https://www.securitron.com/products/m62"),
    ("Securitron M82 Magnetic Lock (1200lb)", "ea", 445, 625, "M82", "Surface-mount maglock, 1200lb holding force", "https://www.securitron.com/products/m82"),
    ("Securitron M680 Shear Lock (1200lb)", "ea", 895, 1255, "M680", "Concealed shear lock for swing doors", "https://www.securitron.com/products/m680"),

    # ========== EXIT DEVICES (button only, no PIR sensors) ==========
    ("Alarm Controls TS-2 Exit Button (Green)", "ea", 85, 119, "TS-2", "Pushbutton REX w/ green illumination", "https://www.alarmcontrols.com/product/ts-2/"),
    ("Camden CM-30 Narrow-Mullion Exit Button", "ea", 95, 135, "CM-30", "Narrow-mullion stainless REX button", "https://www.camdencontrols.com/products/push-buttons/"),
    ("Camden CM-3300 Heavy-Duty Exit Button", "ea", 125, 175, "CM-3300", "Vandal-resistant stainless REX", "https://www.camdencontrols.com/products/push-buttons/"),
    ("Securitron EEB2 Emergency Break-Glass REX", "ea", 165, 229, "EEB2", "Emergency break-glass exit device", "https://www.securitron.com/products/eeb2"),
    ("Dortronics 5276 Large Mushroom Exit Button", "ea", 135, 189, "5276", "Large mushroom-head REX, stainless", "https://www.dortronics.com/products/5276/"),

    # ========== DOOR POSITION / CONTACT SENSORS ==========
    ("Magnetic Door Contact (recessed, concealed)", "ea", 14, 20, "DC-3200-S", "Recessed 3/8-inch concealed contact, 1-inch gap rating", "https://www.sdcsecurity.com/"),
    ("Surface-Mount Door Contact (SPDT)", "ea", 12, 17, "GE 1076C", "Surface-mount 2-screw contact, pigtail leads", "https://www.interlogix.com/"),
    ("Overhead Door Contact (garage/commercial)", "ea", 35, 49, "GRI 4701", "Armored overhead-door contact, 2-inch gap", "https://www.grisk.com/"),

    # ========== STRIKE POWER SUPPLIES ==========
    ("Altronix AL400ULX Power Supply (12/24VDC 4A)", "ea", 195, 275, "AL400ULX", "Dual-voltage access PSU, battery backup, 4A", "https://www.altronix.com/products/AL400ULX"),
    ("Altronix AL600ULX Power Supply (12/24VDC 6A)", "ea", 245, 345, "AL600ULX", "Dual-voltage access PSU, battery backup, 6A", "https://www.altronix.com/products/AL600ULX"),

    # ========== INTERCOMS ==========
    ("Aiphone IX-DV SIP Video Door Station", "ea", 795, 1115, "IX-DV", "IP/SIP vandal-resistant video entry, flush", "https://www.aiphone.com/ix-dv"),
    ("Aiphone IX-DVF-2RA ADA Door Station", "ea", 995, 1395, "IX-DVF-2RA", "ADA-compliant flush IP video station", "https://www.aiphone.com/ix-dvf-2ra"),
    ("Aiphone IX-MV7-HB Master Station (7\")", "ea", 1425, 1995, "IX-MV7-HB", "7\" IP master station w/ handset, touchscreen", "https://www.aiphone.com/ix-mv7-hb"),
    ("Aiphone IX-MV7-B Master Station (no handset)", "ea", 1295, 1815, "IX-MV7-B", "7\" hands-free IP master station", "https://www.aiphone.com/ix-mv7-b"),
    ("Aiphone IX-SSA SIP Sub Station", "ea", 495, 695, "IX-SSA", "Flush-mount SIP audio sub station", "https://www.aiphone.com/ix-ssa"),
    ("2N IP Force Intercom", "ea", 1495, 2095, "9151101C", "Industrial IP intercom, 2MP cam, PoE", "https://www.2n.com/en-US/products/intercoms/ip-force"),
    ("2N IP Verso Intercom", "ea", 995, 1395, "9155101C", "Modular IP intercom, expandable modules", "https://www.2n.com/en-US/products/intercoms/ip-verso"),
    ("Viking E-1600-IP-EWP Emergency Phone", "ea", 895, 1255, "E-1600-IP-EWP", "IP emergency phone, weatherproof", "https://www.vikingelectronics.com/products/e-1600-ip-ewp"),

    # ========== IP CAMERAS — AXIS ==========
    ("Axis M3215-LVE Outdoor Dome", "ea", 595, 835, "02370-001", "2MP outdoor dome, 106° lens, IR", "https://www.axis.com/products/axis-m3215-lve"),
    ("Axis M3216-LVE Indoor/Outdoor Dome", "ea", 795, 1115, "02371-001", "4MP outdoor dome, ideal corridor view", "https://www.axis.com/products/axis-m3216-lve"),
    ("Axis P3265-LVE Outdoor Dome", "ea", 1195, 1675, "02329-001", "2MP outdoor dome, Lightfinder 2.0", "https://www.axis.com/products/axis-p3265-lve"),
    ("Axis P3265-LVE-3 License Plate Dome", "ea", 1895, 2655, "02330-001", "2MP outdoor LPR-optimized dome", "https://www.axis.com/products/axis-p3265-lve"),
    ("Axis Q6135-LE PTZ", "ea", 3995, 5595, "01715-004", "PTZ w/ 32x zoom, laser focus, outdoor", "https://www.axis.com/products/axis-q6135-le"),
    ("Axis P1455-LE Outdoor Bullet", "ea", 1095, 1535, "02095-001", "2MP outdoor bullet, Zipstream, IR", "https://www.axis.com/products/axis-p1455-le"),
    ("Axis M4308-PLE Panoramic", "ea", 1795, 2515, "02273-001", "12MP fisheye 360° panoramic cam", "https://www.axis.com/products/axis-m4308-ple"),

    # ========== IP CAMERAS — HANWHA ==========
    ("Hanwha QNV-C8011 4MP Bullet", "ea", 295, 415, "QNV-C8011", "4MP outdoor bullet, IR, H.265", "https://www.hanwhavisionamerica.com/products/qnv-c8011/"),
    ("Hanwha PNV-A9081R 4K AI Bullet", "ea", 1450, 2025, "PNV-A9081R", "4K outdoor AI bullet, license plate + people", "https://www.hanwhavisionamerica.com/products/pnv-a9081r/"),
    ("Hanwha PNO-A9081R 4K AI Box", "ea", 1395, 1955, "PNO-A9081R", "4K AI box cam (lens sold separately)", "https://www.hanwhavisionamerica.com/products/pno-a9081r/"),
    ("Hanwha XNP-C7310RA AI PTZ", "ea", 3495, 4895, "XNP-C7310RA", "4MP AI PTZ w/ 31x optical zoom, IR 300'", "https://www.hanwhavisionamerica.com/products/xnp-c7310ra/"),

    # ========== IP CAMERAS — AVIGILON ==========
    ("Avigilon H5A-B3 4K Bullet", "ea", 1895, 2655, "4.0C-H5A-BO3-IR", "4K outdoor bullet, AI analytics, Self-Learning", "https://www.avigilon.com/products/video-security/"),
    ("Avigilon H5A-DP3 4K Dome", "ea", 1695, 2375, "4.0C-H5A-DP3", "4K pendant dome, AI analytics", "https://www.avigilon.com/products/video-security/"),

    # ========== VMS / NVR LICENSES ==========
    ("Milestone XProtect Express+ Device License", "ea", 95, 135, "MXPEXPLDL", "Per-camera license, Express+ VMS", "https://www.milestonesys.com/solutions/platform/xprotect-express-plus/"),
    ("Milestone XProtect Professional+ Device License", "ea", 195, 275, "MXPPROPLDL", "Per-camera license, Professional+ VMS", "https://www.milestonesys.com/solutions/platform/xprotect-professional-plus/"),
    ("Genetec Synergis Cloud Link Door License", "ea", 295, 415, "GSC-DL", "Per-door Synergis cloud link license", "https://www.genetec.com/products/unified-security/synergis"),
    ("Genetec Omnicast Standard Camera License", "ea", 195, 275, "GON-CAM-STD", "Per-camera Omnicast Standard license", "https://www.genetec.com/products/unified-security/omnicast"),
    ("Hanwha WAVE VMS Device License", "ea", 149, 209, "WRT-P-CL01", "Per-camera WAVE VMS license, perpetual", "https://www.hanwhavisionamerica.com/products/wisenet-wave/"),
    ("Avigilon ACC 7 Standard Device License", "ea", 195, 275, "ACC7-STD-DVC", "Per-camera ACC 7 Standard license", "https://www.avigilon.com/products/video-security/avigilon-control-center/"),

    # ========== RECORDING SERVERS ==========
    ("Axis S3008 Mk II Recorder (8-ch, 8TB)", "ea", 2195, 3075, "02403-004", "Compact 8-ch recorder w/ integrated PoE", "https://www.axis.com/products/axis-s3008-mk-ii"),
    ("Milestone Husky X2 8-Channel Recorder", "ea", 3495, 4895, "HX2-M10W20N10C8HW", "Pre-loaded XProtect, 8-ch, 10TB", "https://www.milestonesys.com/solutions/platform/husky/"),
    ("Hanwha WAVE WRN-1610S 16-bay NVR", "ea", 3995, 5595, "WRN-1610S", "16-bay WAVE-ready recorder, no HDDs", "https://www.hanwhavisionamerica.com/products/wrn-1610s/"),

    # ========== VIDEO CONFERENCING ==========
    ("Logitech Rally Bar", "ea", 3495, 4895, "960-001309", "All-in-one video bar for medium rooms", "https://www.logitech.com/en-us/products/video-conferencing/conference-cameras/rally-bar.960-001309.html"),
    ("Logitech Rally Bar Mini", "ea", 1999, 2795, "960-001350", "All-in-one video bar for small/huddle rooms", "https://www.logitech.com/en-us/products/video-conferencing/conference-cameras/rally-bar-mini.960-001350.html"),
    ("Logitech Tap Controller", "ea", 895, 1255, "939-001950", "Touch controller for Rally systems", "https://www.logitech.com/en-us/products/video-conferencing/room-solutions/tap.939-001950.html"),
    ("Logitech Rally Board 65", "ea", 5995, 8395, "960-001534", "All-in-one 65\" video collab display", "https://www.logitech.com/en-us/products/video-conferencing/room-solutions/rally-board-65.960-001534.html"),
    ("Poly Studio X50 Video Bar", "ea", 2495, 3495, "2200-86270-001", "All-in-one 4K video bar for small/medium rooms", "https://www.hp.com/us-en/poly/video-conferencing/personal-video/studio-x50.html"),
    ("Poly Studio X70 Video Bar", "ea", 5495, 7695, "2200-87090-001", "Premium dual-cam 4K video bar for large rooms", "https://www.hp.com/us-en/poly/video-conferencing/room-video/studio-x70.html"),
    ("Poly TC8 Touch Controller", "ea", 595, 835, "2200-30760-001", "Touch controller for Poly Studio systems", "https://www.hp.com/us-en/poly/video-conferencing/room-video/tc8.html"),
    ("Neat Bar Gen 2", "ea", 4895, 6855, "NEAT-BAR-2", "All-in-one video bar w/ AI framing", "https://neat.no/bar/"),
    ("Neat Bar Pro", "ea", 5850, 8192, "NEAT-BAR-PRO", "Premium 4K video bar for large rooms", "https://neat.no/bar-pro/"),
    ("Neat Center", "ea", 2290, 3208, "NEAT-CENTER", "AI audio processor + 360° cam", "https://neat.no/center/"),
    ("Neat Pad Touch Controller", "ea", 820, 1150, "NEAT-PAD", "Touch controller for Neat room devices", "https://neat.no/pad/"),

    # ========== VIDEO CONFERENCING — Conference phones ==========
    ("Poly Trio C60 Conference Phone", "ea", 799, 1125, "2200-86590-001", "Premium Teams/Zoom IP conference phone", "https://www.hp.com/us-en/poly/phones/trio/trio-c60.html"),
    ("Poly Trio 8800 Conference Phone (legacy)", "ea", 799, 1125, "2200-66070-019", "IP conference phone, Skype/Teams certified", "https://www.hp.com/us-en/poly/phones/trio/trio-8800.html"),
    ("Poly Trio Expansion Microphone Kit", "pair", 495, 695, "2200-65790-001", "Extension mic pair for Trio phones", "https://www.hp.com/us-en/poly/phones/trio/accessories.html"),

    # ========== AUDIO / DSP ==========
    ("Shure MXA920 Ceiling Array Mic (Round)", "ea", 2895, 4055, "MXA920R-A", "Ceiling array mic, Auto Coverage, IntelliMix DSP", "https://www.shure.com/en-US/products/microphones/mxa920"),
    ("Shure MXA310 Table Array Mic", "ea", 1995, 2795, "MXA310-A", "Tabletop array mic, 4 steerable lobes", "https://www.shure.com/en-US/products/microphones/mxa310"),
    ("Shure IntelliMix P300 DSP", "ea", 2295, 3215, "P300-IMX", "Conferencing DSP, AEC + automixer + AGC", "https://www.shure.com/en-US/products/mixers/p300"),
    ("Biamp TesiraFORTÉ X 400 DSP", "ea", 2895, 4055, "TesiraFORTE-X-400", "AVB/Dante DSP, 4×4 analog I/O, conferencing", "https://www.biamp.com/products/category/digital-signal-processors/tesiraforte-x-400"),
    ("Biamp TesiraFORTÉ X 800 DSP", "ea", 3995, 5595, "TesiraFORTE-X-800", "AVB/Dante DSP, 8×8 analog I/O, conferencing", "https://www.biamp.com/products/category/digital-signal-processors/tesiraforte-x-800"),

    # ========== AUDIO — Speakers & Amps ==========
    ("Atlas FAP62T Ceiling Speaker (6.5\" 70V)", "ea", 135, 189, "FAP62T", "6.5\" 70V ceiling speaker, steel grille", "https://www.atlasied.com/fap62t"),
    ("Atlas FAP82T Ceiling Speaker (8\" 70V)", "ea", 195, 275, "FAP82T", "8\" 70V ceiling speaker w/ backcan", "https://www.atlasied.com/fap82t"),
    ("JBL Control 26CT Ceiling Speaker (pair)", "pair", 225, 315, "CONTROL-26CT-LS", "6.5\" 70V ceiling speaker, white, pair", "https://jblpro.com/products/control-26ct-ls"),
    ("QSC AD-C6T Ceiling Speaker (6.5\" 70V)", "ea", 195, 275, "AD-C6T", "6.5\" 70V ceiling speaker, black/white", "https://www.qsc.com/systems/products/loudspeakers/installed-loudspeakers/ceiling-loudspeakers/ad-c6t/"),
    ("Atlas AA120M Mixer Amplifier (120W)", "ea", 595, 835, "AA120M", "6-input 70V mixer amp, 120W", "https://www.atlasied.com/aa120m"),
    ("Crown CDi 1000 DriveCore Amp (70V)", "ea", 1095, 1535, "CDI-1000", "2-channel 500W/ch Dante-capable amp", "https://www.crownaudio.com/en/products/cdi-1000"),

    # ========== COMMERCIAL DISPLAYS ==========
    ("Samsung QM55C 55\" Commercial Display", "ea", 1195, 1675, "LH55QMCEBGCXGO", "55\" 4K UHD 500nit, 24/7 operation, Tizen", "https://www.samsung.com/us/business/displays/4k-uhd/qm-series/"),
    ("Samsung QM65C 65\" Commercial Display", "ea", 1495, 2095, "LH65QMCEBGCXGO", "65\" 4K UHD 500nit, 24/7 operation, Tizen", "https://www.samsung.com/us/business/displays/4k-uhd/qm-series/"),
    ("Samsung QM75C 75\" Commercial Display", "ea", 2295, 3215, "LH75QMCEBGCXGO", "75\" 4K UHD 500nit, 24/7 operation, Tizen", "https://www.samsung.com/us/business/displays/4k-uhd/qm-series/"),
    ("LG 75UH5N Commercial Display", "ea", 1895, 2655, "75UH5N-E", "75\" 4K 500nit, 24/7, webOS Signage", "https://www.lg.com/us/business/commercial-display/all-commercial-tvs/lg-75uh5n-e"),
    ("NEC P554Q 55\" Commercial Display", "ea", 1795, 2515, "P554Q", "55\" 4K 700nit, 24/7 w/ SpectraView certified", "https://www.sharpnecdisplays.us/products/displays/p554q"),

    # ========== CONTROL SYSTEMS ==========
    ("Crestron HZ-KPCN-B Horizon Keypad (Black)", "ea", 225, 315, "HZ-KPCN-B", "Cresnet horizon keypad, 6-button customizable", "https://www.crestron.com/Products/Lighting-Environmental/Keypads/Cresnet-Keypads/HZ-KPCN-B"),
    ("Crestron TSW-770 7\" Touch Screen", "ea", 1295, 1815, "TSW-770-B-S", "7\" capacitive touchscreen, AirMedia ready", "https://www.crestron.com/Products/Touch-Screens-Computers/Tabletop-Touch-Screens/TSW-770-Series/TSW-770-B-S"),
    ("Extron MLC 226 IP Plus Control Processor", "ea", 1495, 2095, "MLC-226-IP-PLUS", "MediaLink room controller, IR/RS-232/IP", "https://www.extron.com/product/mlc226ipplus"),
    ("Control4 C4-KC8-Z 8-Button Keypad", "ea", 195, 275, "C4-KC8-Z", "Wireless 8-button keypad, engraving customizable", "https://www.control4.com/products/keypads"),
    ("RTI KX2 Touchscreen Controller", "ea", 695, 975, "KX2", "Wall-mount in-room controller", "https://www.rticorp.com/products/kx2"),

    # ========== HDMI-over-IP / Extenders ==========
    ("Crestron DM NVX E30 Encoder", "ea", 1595, 2235, "DM-NVX-E30", "4K60 4:4:4 HDR AV-over-IP encoder", "https://www.crestron.com/Products/Catalog/AV-Over-IP/DM-NVX-AV-Over-IP/Video-Endpoint/DM-NVX-E30"),
    ("Crestron DM NVX D30 Decoder", "ea", 1595, 2235, "DM-NVX-D30", "4K60 4:4:4 HDR AV-over-IP decoder", "https://www.crestron.com/Products/Catalog/AV-Over-IP/DM-NVX-AV-Over-IP/Video-Endpoint/DM-NVX-D30"),
    ("Crestron DM NVX E20 Wallplate Encoder", "ea", 995, 1395, "DM-NVX-E20-2G", "4K60 4:2:0 wallplate AV-over-IP encoder", "https://www.crestron.com/Products/Catalog/AV-Over-IP/DM-NVX-AV-Over-IP/Video-Endpoint/DM-NVX-E20"),
    ("Crestron DM NVX 351 Encoder/Decoder", "ea", 2995, 4195, "DM-NVX-351", "4K60 4:4:4 HDR Network AV transceiver", "https://www.crestron.com/Products/Catalog/AV-Over-IP/DM-NVX-AV-Over-IP/Video-Endpoint/DM-NVX-351"),
    ("AVPro Edge AC-EX100-UHD HDBaseT Extender", "pair", 695, 975, "AC-EX100-UHD-KIT", "HDBaseT TX/RX kit, 4K60 to 330'", "https://www.avproedge.com/AC-EX100-UHD-KIT"),

    # ========== ENTERPRISE INTERCOM (broadcast/production) ==========
    ("Clear-Com MS-702 Encore 2-Channel Main Station", "ea", 1595, 2235, "MS-702", "1RU 2-channel analog partyline main station", "https://www.clearcom.com/Product/MS-702"),
    ("Clear-Com HelixNet HMS-4X Digital Main Station", "ea", 3495, 4895, "HMS-4X", "Digital partyline master station, 4-channel", "https://www.clearcom.com/Product/HMS-4X"),
    ("Clear-Com HelixNet HXII-BP-X4 Beltpack", "ea", 795, 1115, "HXII-BP-X4", "4-channel digital partyline beltpack", "https://www.clearcom.com/Product/HXII-BP-X4"),
    ("Clear-Com LQ-2W2 Portable IP Interface", "ea", 1895, 2655, "LQ-2W2", "2-channel partyline-over-IP gateway", "https://www.clearcom.com/Product/LQ-2W2"),

    # ========== RACK POWER (PDU / UPS) ==========
    ("APC AP9571A Basic Rack PDU 15A", "ea", 195, 275, "AP9571A", "15A 120V 1U basic rack PDU, 8 outlets", "https://www.apc.com/us/en/product/AP9571A"),
    ("Tripp Lite PDUMH15-NET Metered PDU", "ea", 495, 695, "PDUMH15-NET", "15A 120V metered network PDU, 14 outlets", "https://www.tripplite.com/15a-120v-metered-pdu-2u-rack-mount"),
    ("APC Smart-UPS 1500VA RM 2U", "ea", 895, 1255, "SMT1500RM2UC", "1500VA/1000W rack UPS w/ Network Mgmt", "https://www.apc.com/us/en/product/SMT1500RM2UC"),
    ("APC Smart-UPS 3000VA RM 2U", "ea", 1895, 2655, "SMT3000RM2UC", "3000VA/2700W rack UPS w/ Network Mgmt", "https://www.apc.com/us/en/product/SMT3000RM2UC"),
    ("CyberPower OR1500LCDRM2U UPS", "ea", 300, 421, "OR1500LCDRM2U", "1500VA/900W 2U line-interactive rack UPS", "https://www.cyberpowersystems.com/product/ups/pfc-sinewave/or1500lcdrm2u/"),

    # ========== GENERAL RACK HARDWARE ==========
    ("42U Server Rack (Tripp Lite SR42UB)", "ea", 895, 1255, "SR42UB", "42U 4-post server rack, black, 600x1070mm", "https://www.tripplite.com/42u-smartrack-premium-enclosure-server-rack"),
    ("12U Wall-Mount Rack (Tripp Lite SRW12USG)", "ea", 395, 555, "SRW12USG", "12U hinged wall rack, glass door, 20\" deep", "https://www.tripplite.com/12u-smartrack-low-profile-switch-depth-wall-mount-rack-enclosure"),
    ("6U Wall-Mount Rack (Tripp Lite SRW6U)", "ea", 245, 345, "SRW6U", "6U hinged wall rack, glass door", "https://www.tripplite.com/6u-smartrack-switch-depth-wall-mount-rack-enclosure"),
    ("Panduit PE2U Horizontal Cable Manager (2U)", "ea", 85, 119, "NM2", "2U horizontal finger manager, front & rear", "https://www.panduit.com/en/products/cable-management.html"),

    # ========== PATCH PANELS / KEYSTONES ==========
    ("Cat6A 24-Port Patch Panel (shielded)", "ea", 195, 275, "NKPP24Y-6A", "Panduit Cat6A 24-port shielded patch panel, 1U", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A 48-Port Patch Panel (shielded)", "ea", 395, 555, "NKPP48Y-6A", "Panduit Cat6A 48-port shielded patch panel, 2U", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Keystone Jack (shielded, white)", "ea", 8, 12, "NKP6X88MYL", "Panduit Cat6A keystone jack, 10GbE rated", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Keystone Jack (shielded, blue)", "ea", 8, 12, "NKP6X88MBU", "Panduit Cat6A keystone jack, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Keystone Jack (shielded, green)", "ea", 8, 12, "NKP6X88MGR", "Panduit Cat6A keystone jack, green (security/camera)", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),

    # ========== FACEPLATES ==========
    ("1-Port Stainless Steel Faceplate", "ea", 4, 6, "CFP1SY", "Panduit 1-port SS faceplate", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("2-Port Stainless Steel Faceplate", "ea", 5, 7, "CFP2SY", "Panduit 2-port SS faceplate", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("4-Port Stainless Steel Faceplate", "ea", 7, 10, "CFP4SY", "Panduit 4-port SS faceplate", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("6-Port Stainless Steel Faceplate", "ea", 9, 13, "CFP6SY", "Panduit 6-port SS faceplate", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),

    # ========== WAP MOUNTS / PoE INJECTORS ==========
    ("Oberon 1063-00 Drop-Ceiling WAP Mount", "ea", 65, 95, "1063-00", "Above-ceiling tile WAP enclosure w/ 2x2 tile", "https://oberoninc.com/products/1063-00/"),
    ("Oberon 1064-00 Hard-Surface WAP Mount", "ea", 85, 119, "1064-00", "Hard-surface (drywall) WAP mounting plate", "https://oberoninc.com/products/1064-00/"),
    ("TP-Link TL-POE160S PoE++ Injector (60W)", "ea", 55, 79, "TL-POE160S", "Gigabit 802.3bt PoE++ injector, 60W", "https://www.tp-link.com/us/business-networking/accessory/tl-poe160s/"),

    # ========== UNIFI ACCESSORIES ==========
    ("UniFi U-POE-In Adapter", "ea", 15, 22, "U-POE-In", "PoE splitter adapter for UniFi devices", "https://store.ui.com/us/en/products/u-poe-in"),
    ("UniFi Smart Power Plug", "ea", 29, 42, "USP-Plug", "Zigbee smart plug, remote power control", "https://store.ui.com/us/en/products/usp-plug"),
]

# ============================================================================
# COMMERCIAL MATERIALS (Sheet 2) — same schema as Equipment
# ============================================================================
commercial_materials = [
    # ========== CAT6A CABLE BULK ==========
    ("Cat6A Plenum Cable (blue)", "1000 ft", 425, 595, "Belden 10GXS12", "Cat6A 23AWG plenum UTP, blue (data)", "https://catalog.belden.com/index.cfm?event=pd&p=PF_10GXS12"),
    ("Cat6A Plenum Cable (green)", "1000 ft", 425, 595, "Belden 10GXS12-GRN", "Cat6A 23AWG plenum UTP, green (security/camera)", "https://catalog.belden.com/index.cfm?event=pd&p=PF_10GXS12"),
    ("Cat6A Plenum Cable (yellow)", "1000 ft", 425, 595, "Belden 10GXS12-YEL", "Cat6A 23AWG plenum UTP, yellow (AV)", "https://catalog.belden.com/index.cfm?event=pd&p=PF_10GXS12"),
    ("Cat6A Riser Cable (blue)", "1000 ft", 325, 455, "Belden 10GXW12", "Cat6A 23AWG riser UTP, blue", "https://catalog.belden.com/index.cfm?event=pd&p=PF_10GXW12"),
    ("Cat6A Outdoor Direct-Burial Cable", "1000 ft", 725, 1015, "Belden 10GXS13", "Cat6A outdoor gel-filled direct-burial", "https://catalog.belden.com/index.cfm?event=pd&p=PF_10GXS13"),

    # ========== CAT6A PATCH CORDS ==========
    ("Cat6A Patch Cord 1ft (blue)", "ea", 5, 7, "UTP28X1BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 3ft (blue)", "ea", 6, 9, "UTP28X3BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 5ft (blue)", "ea", 7, 10, "UTP28X5BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 7ft (blue)", "ea", 8, 12, "UTP28X7BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 10ft (blue)", "ea", 10, 14, "UTP28X10BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 14ft (blue)", "ea", 13, 18, "UTP28X14BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 25ft (blue)", "ea", 18, 25, "UTP28X25BU", "Panduit Cat6A 28AWG patch cord, blue", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 3ft (green)", "ea", 6, 9, "UTP28X3GR", "Panduit Cat6A 28AWG patch cord, green", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Cat6A Patch Cord 7ft (green)", "ea", 8, 12, "UTP28X7GR", "Panduit Cat6A 28AWG patch cord, green", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),

    # ========== FIBER CABLE ==========
    ("OS2 Single-Mode Fiber 12-strand Plenum", "1000 ft", 695, 975, "Corning AW2012", "12-strand OS2 SMF plenum, distribution cable", "https://www.corning.com/optical-communications/worldwide/en/home.html"),
    ("OM4 Multi-Mode Fiber 12-strand Plenum", "1000 ft", 895, 1255, "Corning AE2212", "12-strand OM4 MMF plenum, 50/125 micron", "https://www.corning.com/optical-communications/worldwide/en/home.html"),
    ("OS2 Single-Mode Fiber 24-strand Plenum", "1000 ft", 1195, 1675, "Corning AW2024", "24-strand OS2 SMF plenum", "https://www.corning.com/optical-communications/worldwide/en/home.html"),

    # ========== FIBER PATCH CORDS ==========
    ("LC-LC OS2 Duplex Patch Cord 1m", "ea", 12, 17, "FJLCOS2-01", "Single-mode duplex LC-LC, 1 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC-LC OS2 Duplex Patch Cord 2m", "ea", 14, 20, "FJLCOS2-02", "Single-mode duplex LC-LC, 2 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC-LC OS2 Duplex Patch Cord 3m", "ea", 16, 22, "FJLCOS2-03", "Single-mode duplex LC-LC, 3 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC-LC OM4 Duplex Patch Cord 1m", "ea", 15, 21, "FJLCOM4-01", "Multi-mode OM4 duplex LC-LC, 1 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC-LC OM4 Duplex Patch Cord 3m", "ea", 22, 31, "FJLCOM4-03", "Multi-mode OM4 duplex LC-LC, 3 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("SC-LC OM4 Duplex Patch Cord 3m", "ea", 24, 34, "FJSCLCOM4-03", "Multi-mode OM4 duplex SC-LC, 3 meter", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),

    # ========== FIBER HARDWARE ==========
    ("LC Adapter Panel 12-Port OS2", "ea", 65, 92, "FAP12WBUDLCZ", "12-port LC duplex adapter panel, SM blue", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC Adapter Panel 24-Port OS2", "ea", 95, 135, "FAP24WBUDLCZ", "24-port LC duplex adapter panel, SM blue", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("Wall-Mount Fiber Enclosure (6-adapter)", "ea", 295, 415, "FWME2", "Panduit 6-port wall-mount fiber enclosure", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("MPO Cassette OM4 12-Fiber LC", "ea", 395, 555, "FMC1M-12X12-F1", "OM4 MPO-12 to LC cassette, factory terminated", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),
    ("LC Pigtail OS2 Single-Mode (1.5m)", "ea", 22, 31, "FOTP7Y-01.5", "Single-mode LC pigtail, 900um jacket, 1.5m", "https://www.panduit.com/en/products/network-cabling-system/fiber-connectivity.html"),

    # ========== PATHWAY ==========
    ("Cable Tray 12\" (10 ft section)", "ea", 195, 275, "WBT-12-240", "Wire basket cable tray, 12\" wide, 10ft", "https://www.chatsworth.com/en-us/products/cable-management/cable-tray"),
    ("J-Hook 2\" with beam clamp", "ea", 3, 4, "CM16BC", "Caddy 2\" J-hook w/ beam clamp, 16ga steel", "https://www.erico.com/products/cable-support"),
    ("J-Hook 4\" with beam clamp", "ea", 5, 7, "CM32BC", "Caddy 4\" J-hook w/ beam clamp, 16ga steel", "https://www.erico.com/products/cable-support"),
    ("EMT Conduit 1\" (10 ft)", "ea", 15, 21, "EMT-1-10", "1\" EMT conduit, 10 ft length", "https://www.allied-electronics.com/"),
    ("EMT Conduit 3/4\" (10 ft)", "ea", 10, 14, "EMT-075-10", "3/4\" EMT conduit, 10 ft length", "https://www.allied-electronics.com/"),

    # ========== FIRESTOPPING ==========
    ("3M CP 25WB+ Firestop Sealant (tube)", "ea", 25, 35, "CP-25WB+", "Caulk-grade firestop sealant, UL-listed", "https://www.3m.com/3M/en_US/p/d/b5005181040/"),
    ("3M FireDam Spray 200 (gal)", "ea", 145, 205, "FS-200", "Sprayable firestop for large openings", "https://www.3m.com/3M/en_US/p/d/b5005181047/"),
    ("Hilti CP 643N Firestop Collar 2\"", "ea", 35, 49, "CP-643N-2", "Intumescent firestop collar, 2\" pipe", "https://www.hilti.com/firestop-products"),

    # ========== LABELING ==========
    ("Brady M611 Label Printer Cartridge", "ea", 65, 92, "M611-R6000", "Handheld thermal cartridge, wire marking", "https://www.bradyid.com/label-makers"),
    ("Panduit P1 HellermannTyton Cable Labels", "roll", 45, 63, "T150X15VM-BK", "Self-laminating vinyl cable labels, 500/roll", "https://www.hellermann.tyton.com/"),

    # ========== TERMINATION / TOOLS CONSUMABLES ==========
    ("RJ45 Plugs Cat6A Shielded (100-pack)", "pack", 45, 63, "SP688-C", "Panduit shielded RJ45 plugs for Cat6A, 100pk", "https://www.panduit.com/en/products/network-cabling-system/copper-connectivity.html"),
    ("Velcro Cable Ties (8\" roll of 900)", "roll", 65, 92, "HLT-8PK-900", "Reusable hook-and-loop cable ties, black", "https://www.velcro.com/commercial/"),
    ("Zip Ties Assortment (1000-pack)", "pack", 35, 49, "ZT-1000-MIX", "Nylon zip ties, 4\"/6\"/8\" mixed", "https://www.uline.com/"),

    # ========== RACK ACCESSORIES ==========
    ("Horizontal Cable Manager 1U (finger type)", "ea", 65, 92, "WMPF1E", "Panduit 1U horizontal finger manager", "https://www.panduit.com/en/products/cable-management.html"),
    ("Horizontal Cable Manager 2U (finger type)", "ea", 95, 135, "WMPF2E", "Panduit 2U horizontal finger manager", "https://www.panduit.com/en/products/cable-management.html"),
    ("Blanking Panel 1U (black plastic)", "ea", 8, 12, "CPP1U-BK", "Tool-less snap-in 1U blanking panel", "https://www.panduit.com/en/products/network-cabling-system.html"),
    ("Blanking Panel 2U (black metal)", "ea", 15, 21, "CPP2U-SA", "Screw-mount 2U metal blanking panel", "https://www.panduit.com/en/products/network-cabling-system.html"),
    ("1U Cantilever Shelf", "ea", 65, 92, "1915-1-001-01", "Middle Atlantic 1U 10\" cantilever rack shelf", "https://www.legrandav.com/"),

    # ========== MDF / IDF HARDWARE ==========
    ("Ground Bar Kit (rack)", "ea", 95, 135, "RGRB19U", "Panduit 19\" rack ground bar w/ lugs", "https://www.panduit.com/en/products/grounding-bonding.html"),
    ("Copper Ground Wire #6 AWG (100 ft)", "ea", 95, 135, "CU-6-SOL", "Solid copper ground wire, 100ft", "https://www.southwire.com/"),
    ("Plywood Backboard 4x8 Fire-Treated", "ea", 95, 135, "PW-FR-48", "Fire-treated 3/4\" plywood telecom backboard", "https://www.homedepot.com/"),
]

# ============================================================================
# COMMERCIAL SERVICES (Sheet 3)
# Format: (name, units, labor_hours, description)
# Price = hours × $63.26, Sale Price = hours × $100.00
# ============================================================================
commercial_services_raw = [
    # ========== DROP INSTALLATION (BICSI-aligned) ==========
    ("Cat6A Data Drop - Complete Install", "drop", 1.00, "Pull + terminate + test + label, ~125 LF run, open plenum"),
    ("Cat6 Data Drop - Complete Install (VE alt only)", "drop", 0.75, "Pull + terminate + test + label, Cat6 (never GCC default)"),
    ("Cat6A Camera Drop - Complete Install", "drop", 1.50, "Drop + camera mount + aim + test, ~150 LF run"),
    ("Cat6A WAP Drop - Complete Install", "drop", 1.25, "Drop + AP mount + ceiling access + test"),
    ("Cat6A ACS Door - Cable Composite (per door)", "door", 2.50, "Reader + REX + EOL + contacts + reader prep"),
    ("Security Camera Install (full mount + aim + configure)", "cam", 2.25, "Mount + aim + config + verify w/ GC"),
    ("Intercom Station Install", "station", 1.50, "Mount + wire + program station on intercom system"),
    ("Ceiling Speaker Install (70V)", "speaker", 0.75, "Locate + cut + pull + terminate + level set"),
    ("RG6 TV Drop - Complete Install", "drop", 0.65, "Coax pull + fitting + test, simpler than Cat6"),
    ("HDMI Active Drop - Complete Install", "drop", 1.00, "Fiber HDMI pull + test at both ends"),

    # ========== TERMINATION / CERT ==========
    ("Copper Port Termination + Dress (Cat6/6A)", "port", 0.083, "5 min/port BICSI TDMM — punch, dress, verify"),
    ("Fluke Copper Permanent Link Certification", "port", 0.083, "5 min/port — test, save result, label"),
    ("Fiber Fusion Splice", "strand", 0.25, "15 min/strand — cleave + splice + heat-shrink + sleeve"),
    ("Fiber Mechanical Splice", "strand", 0.167, "10 min/strand — cleave + mechanical splice"),
    ("Fiber OTDR Certification (bidirectional)", "run", 0.50, "30 min/run — shoot both ends + save trace"),
    ("Coax Sweep Certification", "run", 0.25, "15 min/run — TDR + loss test"),

    # ========== RACK / INFRASTRUCTURE ==========
    ("42U Rack Assembly + Mount + Ground", "rack", 8.00, "Assemble + anchor + ground bond + shelves"),
    ("MDF Full Buildout", "mdf", 20.00, "Rack + patch panels + fiber housings + ground bar + initial dress"),
    ("IDF Full Buildout (12U wall)", "idf", 8.00, "Wall rack + 1x 48-port patch + ground + backboard prep"),
    ("Rack Dress-Out Only (existing infrastructure)", "rack", 8.00, "Patch in + organize + label + verify, no structural"),
    ("Cable Tray Install (per LF)", "LF", 0.25, "Pathway section install, incl. hardware"),
    ("J-Hook Install (per each)", "ea", 0.083, "5 min — locate + install w/ beam clamp"),
    ("Firestop per Penetration", "pen", 0.33, "20 min — includes material application + cure monitoring"),
    ("Cable Labeling + Documentation (per port)", "port", 0.067, "4 min/port — print + verify + record in as-built"),

    # ========== PROJECT MANAGEMENT / MOBILIZATION ==========
    ("Mobilization — Commercial TI", "flat", 8.00, "Initial setup, site safety orientation, tool stage"),
    ("Demobilization + Punch Walk", "flat", 6.00, "Final cleanup + punch list review + tool demob"),
    ("Project Management (per week)", "week", 4.00, "Weekly GC coordination, RFI response, sched update"),
    ("GC-Mandated Daily Cleanup", "day", 0.50, "End-of-day work area cleanup to GC standard"),

    # ========== DESIGN / CONSULTING ==========
    ("Site Survey / Pre-Construction Walk", "hr", 1.00, "Walk-through survey by foreman, pre-bid/pre-mobe"),
    ("RCDD Review & Stamp", "flat", 25.00, "BICSI RCDD-stamped design review (when spec requires)"),
    ("ICRA Compliance Package (healthcare)", "flat", 25.00, "Infection Control Risk Assessment docs + badging + training"),
    ("BICSI Installer Certified Labor Adder", "hr", 0.10, "+10% rate adder when spec requires certified install"),

    # ========== COMMISSIONING & CLOSEOUT ==========
    ("Submittals Package (Shop Drawings + Product Data)", "each", 0.50, "Per-submittal prep + transmittal + GC upload"),
    ("Closeout Documentation Package", "pkg", 7.50, "As-builts + cert reports + O&M + warranty registration"),
    ("Manufacturer Training (Client Turnover)", "hr", 1.00, "End-user training on installed systems"),

    # ========== POST-OCCUPANCY ==========
    ("MAC Post-Occupancy (per drop + $100/hr)", "drop", 1.00, "Post-occupancy moves/adds/changes base rate"),
]

# ============================================================================
# RESIDENTIAL EQUIPMENT, MATERIALS, SERVICES — preserved from v3
# ============================================================================
residential_equipment = [
    ("Residential Router (Standard)", "ea", 85, 125, "Netgear R6900", "AC1900 dual-band router, home use", "https://www.netgear.com/home/wifi/routers/r6900/"),
    ("Residential WiFi Mesh 3-Pack", "pack", 295, 415, "TP-Link Deco X55", "WiFi 6 mesh system, 3-unit", "https://www.tp-link.com/us/home-networking/deco/deco-x55/"),
    ("Residential 8-Port Gigabit Switch", "ea", 35, 49, "TP-Link TL-SG108", "Unmanaged 8-port gigabit switch", "https://www.tp-link.com/us/business-networking/soho-switch/tl-sg108/"),
    ("Structured Media Panel (30\")", "ea", 195, 275, "Leviton 49605-30W", "30\" structured media enclosure, hinged", "https://www.leviton.com/products/49605-30w"),
    ("Structured Media Panel (42\")", "ea", 245, 345, "Leviton 49605-42W", "42\" structured media enclosure, hinged", "https://www.leviton.com/products/49605-42w"),
    ("6U Open Frame Wall Rack", "ea", 125, 175, "Tripp Lite SRWF6U", "6U open-frame wall rack, low-voltage", "https://www.tripplite.com/6u-low-profile-switch-depth-wall-mount-4-post-open-frame-rack"),
    ("Coax Splitter 2-Way (wall-plate)", "ea", 15, 21, "Holland DPD2", "2-way RG6 wall-plate splitter, 5-1000 MHz", "https://www.hollandelectronics.com/"),
    ("Coax Splitter 4-Way (panel)", "ea", 18, 25, "Holland DPD4", "4-way RG6 panel splitter", "https://www.hollandelectronics.com/"),
    ("Residential Cat6 Patch Panel (12-port)", "ea", 45, 63, "Monoprice 5358", "12-port Cat6 keystone panel for SMP", "https://www.monoprice.com/"),
    ("In-Wall Speaker 6.5\" (pair)", "pair", 125, 175, "Polk RC60i", "6.5\" in-wall speakers, pair, paintable", "https://www.polkaudio.com/en-us/product/rc60i"),
    ("In-Wall Speaker 8\" (pair)", "pair", 195, 275, "Polk RC80i", "8\" in-wall speakers, pair, paintable", "https://www.polkaudio.com/en-us/product/rc80i"),
    ("In-Wall Volume Control", "ea", 45, 63, "OEM VC-100", "Impedance-matching wall volume control", "https://www.monoprice.com/"),
    ("Home Theater Subwoofer (in-wall)", "ea", 495, 695, "Polk RC85i-SUB", "In-wall passive subwoofer, paintable", "https://www.polkaudio.com/en-us/product/"),
    ("LV Bracket 1-Gang (remodel)", "ea", 3, 4, "Carlon SC100A", "1-gang LV bracket, nail-less remodel", "https://www.carlon.com/"),
    ("LV Bracket 2-Gang (remodel)", "ea", 5, 7, "Carlon SC200A", "2-gang LV bracket, nail-less remodel", "https://www.carlon.com/"),
    ("Residential Patch Panel (24-port)", "ea", 85, 125, "Monoprice 5339", "24-port Cat6 keystone panel for SMP", "https://www.monoprice.com/"),
    ("Residential Router (Premium)", "ea", 295, 415, "ASUS RT-AX88U Pro", "WiFi 6 AX6000 premium router", "https://www.asus.com/networking-iot-servers/wifi-routers/asus-wifi-routers/rt-ax88u-pro/"),
]

residential_materials = [
    ("Cat6 Riser Cable (blue)", "1000 ft", 175, 245, "Commscope Cat6-CMR", "Cat6 23AWG riser UTP, blue (residential)", "https://www.commscope.com/"),
    ("Cat6 Riser Cable (white)", "1000 ft", 175, 245, "Commscope Cat6-CMR-WH", "Cat6 23AWG riser UTP, white", "https://www.commscope.com/"),
    ("RG6 Quad-Shield Riser (black)", "1000 ft", 195, 275, "Belden 7915A", "RG6 quad-shield riser-rated, black", "https://catalog.belden.com/"),
    ("16/2 Speaker Wire Riser (white)", "500 ft", 95, 135, "Monoprice 2811", "16AWG 2-conductor speaker wire, CL3R", "https://www.monoprice.com/"),
    ("14/2 Speaker Wire Riser (white)", "500 ft", 125, 175, "Monoprice 2814", "14AWG 2-conductor speaker wire, CL3R", "https://www.monoprice.com/"),
    ("Direct-Burial Cat6 Outdoor", "1000 ft", 395, 555, "Belden 7812ADB", "Gel-filled Cat6 direct-burial, black", "https://catalog.belden.com/"),
    ("12-2 Romex w/ Ground", "250 ft", 165, 231, "Southwire 12-2", "12AWG 2-conductor w/ ground, 250ft roll", "https://www.southwire.com/"),
    ("Cat6 Keystone Jack (white)", "ea", 3, 4, "CJ688TGWH", "Panduit Cat6 keystone jack, white", "https://www.panduit.com/"),
    ("Cat6 Keystone Jack (ivory)", "ea", 3, 4, "CJ688TGIW", "Panduit Cat6 keystone jack, ivory", "https://www.panduit.com/"),
    ("RG6 F-Connector Compression", "ea", 1, 2, "CMP-RG6-QS", "Compression-style F connector, quad-shield", "https://www.hollandelectronics.com/"),
    ("1-Port White ABS Faceplate", "ea", 1, 2, "CFP1WH", "Panduit 1-port ABS faceplate, white", "https://www.panduit.com/"),
    ("2-Port White ABS Faceplate", "ea", 2, 3, "CFP2WH", "Panduit 2-port ABS faceplate, white", "https://www.panduit.com/"),
    ("4-Port White ABS Faceplate", "ea", 3, 4, "CFP4WH", "Panduit 4-port ABS faceplate, white", "https://www.panduit.com/"),
    ("Banana Plug Binding Post (speaker)", "ea", 5, 7, "BP-SPK-1", "Dual banana binding post wall plate", "https://www.monoprice.com/"),
    ("RG6 Coax Barrel Connector", "ea", 1, 2, "F-BARREL", "F-type female-to-female barrel", "https://www.hollandelectronics.com/"),
]

residential_services_raw = [
    ("Cat6 Data Drop (residential)", "drop", 0.75, "Pull + terminate + test, typical home 100 LF run"),
    ("RG6 TV Drop (residential)", "drop", 0.65, "Coax pull + F-fitting + sweep test"),
    ("Speaker Drop (in-wall, 16/2)", "drop", 0.50, "Pull + terminate both ends + dress"),
    ("In-Wall Speaker Install (pair)", "pair", 1.00, "Cut-in + speaker mount + paint prep (no paint)"),
    ("In-Wall Volume Control Install", "ea", 0.50, "Cut-in + wire + test + low-profile cover"),
    ("Structured Media Panel Buildout (30\")", "smp", 3.00, "Mount + pull home runs + terminate + label"),
    ("Structured Media Panel Buildout (42\")", "smp", 4.00, "Larger panel + more home runs + organize"),
    ("Residential WiFi Mesh Install (3-pack)", "sys", 2.00, "Place + configure + test 3 mesh nodes"),
    ("Residential Router Install", "ea", 1.00, "Swap or new install + config + basic network setup"),
    ("Coax Splitter Install", "ea", 0.25, "Wire + label + verify signal"),
    ("Residential PM (small job)", "job", 2.00, "Scheduling + single-family coordination + punch"),
    ("Residential Mobilization (per day)", "day", 1.50, "Home protection + setup + tools stage"),
    ("Doorbell Camera Install (residential)", "ea", 1.00, "Wire + mount + app setup + WiFi connect"),
    ("WiFi Site Survey (residential)", "hr", 1.00, "Walk-through + coverage map + recs"),
    ("Cable Rerun / Patch (residential)", "drop", 1.50, "Re-pull failed drop or add new run in finished space"),
    ("Smart TV Mount + Hide Wires", "tv", 1.25, "Wall mount + in-wall HDMI + power relocation"),
    ("On-Site Service Call (hourly)", "hr", 1.00, "Truck-roll residential service, billed portal-to-portal"),
    ("Residential Closeout Walk", "job", 1.00, "Walk homeowner through system + labels + warranty"),
]


# ============================================================================
# BUILD FUNCTIONS
# ============================================================================
def make_service_row(svc):
    """Expand (name, units, hrs, desc) → (name, units, hrs, price, sale, desc)"""
    name, units, hrs, desc = svc
    price = round(hrs * LABOR_COST_PER_HR, 2)
    sale_price = round(hrs * LABOR_SALE_PER_HR, 2)
    return (name, units, hrs, price, sale_price, desc)


def build_xlsx(path, equipment, materials, services_raw, title):
    """Write 3-sheet XLSX."""
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor=BRAND_GREEN)
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    body_font = Font(name="Calibri", size=10)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    def write_item_sheet(ws, headers, rows, col_widths):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = left if c in (1, 5, 6) else (right if c in (3, 4) else center)
                if c in (3, 4):  # Price cols
                    cell.number_format = '"$"#,##0.00'
                if c == 7 and isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def write_service_sheet(ws, rows):
        headers = ["Service Name", "Units", "Labor Hours", "Price", "Sale Price", "Short Description"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center
        for r, svc in enumerate(rows, start=2):
            row = make_service_row(svc)
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = left if c in (1, 6) else (right if c in (3, 4, 5) else center)
                if c == 3:
                    cell.number_format = '0.000'
                if c in (4, 5):
                    cell.number_format = '"$"#,##0.00'
        for i, w in enumerate([42, 10, 13, 13, 13, 60], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    # Sheet 1: Equipment
    ws1 = wb.active
    ws1.title = "Equipment"
    item_headers = ["Item Name", "Units", "Price", "Sale Price", "Model Name", "Short Description", "Purchase Link"]
    col_widths_items = [52, 10, 12, 12, 28, 60, 60]
    write_item_sheet(ws1, item_headers, equipment, col_widths_items)

    # Sheet 2: Materials
    ws2 = wb.create_sheet("Materials")
    write_item_sheet(ws2, item_headers, materials, col_widths_items)

    # Sheet 3: Services
    ws3 = wb.create_sheet("Services")
    write_service_sheet(ws3, services_raw)

    wb.save(path)
    return path


def build_md(path, equipment, materials, services_raw, title, subtitle):
    """Write flat markdown export."""
    lines = [f"# {title}", "", f"*{subtitle}*", ""]
    lines.append(f"- **Equipment items:** {len(equipment)}")
    lines.append(f"- **Materials items:** {len(materials)}")
    lines.append(f"- **Services items:** {len(services_raw)}")
    lines.append(f"- **Services cost basis:** ${LABOR_COST_PER_HR}/hr (fully-burdened blended internal)")
    lines.append(f"- **Services sale basis:** ${LABOR_SALE_PER_HR:.2f}/hr (client-facing billed)")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Equipment
    lines.append("## Sheet 1 — Equipment")
    lines.append("")
    lines.append("| Item Name | Units | Price | Sale Price | Model | Description | Link |")
    lines.append("|---|---|---:|---:|---|---|---|")
    for row in equipment:
        name, units, price, sale, model, desc, link = row
        lines.append(f"| {name} | {units} | ${price:,.2f} | ${sale:,.2f} | {model} | {desc} | [link]({link}) |")
    lines.append("")

    # Materials
    lines.append("## Sheet 2 — Materials")
    lines.append("")
    lines.append("| Item Name | Units | Price | Sale Price | Model | Description | Link |")
    lines.append("|---|---|---:|---:|---|---|---|")
    for row in materials:
        name, units, price, sale, model, desc, link = row
        lines.append(f"| {name} | {units} | ${price:,.2f} | ${sale:,.2f} | {model} | {desc} | [link]({link}) |")
    lines.append("")

    # Services
    lines.append("## Sheet 3 — Services")
    lines.append("")
    lines.append("| Service Name | Units | Labor Hours | Price | Sale Price | Description |")
    lines.append("|---|---|---:|---:|---:|---|")
    for svc in services_raw:
        name, units, hrs, desc = svc
        price = round(hrs * LABOR_COST_PER_HR, 2)
        sale = round(hrs * LABOR_SALE_PER_HR, 2)
        lines.append(f"| {name} | {units} | {hrs:.3f} | ${price:,.2f} | ${sale:,.2f} | {desc} |")
    lines.append("")

    with open(path, "w") as f:
        f.write("\n".join(lines))
    return path


def main():
    # Build commercial
    comm_xlsx = os.path.join(OUT_DIR, "Master Catalog.xlsx")
    comm_md = os.path.join(OUT_DIR, "Master Catalog.md")
    build_xlsx(comm_xlsx, commercial_equipment, commercial_materials, commercial_services_raw,
               "Master Catalog - Commercial")
    build_md(comm_md, commercial_equipment, commercial_materials, commercial_services_raw,
             "GCC Master Catalog — Commercial",
             "Reference data only. Prices refreshed weekly from supplier web (price + tax + ship). "
             "See System/memory.md §Master Catalog Structure for structure rules.")

    # Build residential
    res_xlsx = os.path.join(OUT_DIR, "Master Catalog - Residential.xlsx")
    res_md = os.path.join(OUT_DIR, "Master Catalog - Residential.md")
    build_xlsx(res_xlsx, residential_equipment, residential_materials, residential_services_raw,
               "Master Catalog - Residential")
    build_md(res_md, residential_equipment, residential_materials, residential_services_raw,
             "GCC Master Catalog — Residential",
             "Reference data only. Residential-specific items; all commercial items live in "
             "the parallel Master Catalog.xlsx.")

    print(f"✓ Commercial:   {len(commercial_equipment)} equipment | {len(commercial_materials)} materials | {len(commercial_services_raw)} services")
    print(f"✓ Residential:  {len(residential_equipment)} equipment | {len(residential_materials)} materials | {len(residential_services_raw)} services")
    print()
    for p in [comm_xlsx, comm_md, res_xlsx, res_md]:
        print(f"  → {p} ({os.path.getsize(p)} bytes)")


if __name__ == "__main__":
    main()
