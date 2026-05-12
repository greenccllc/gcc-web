"""
BTQ Generator — GCC LV Div27-28

Generates the multi-sheet BTQ (Bid Task Quantities) Excel workbook for a GCC
bid. BTQ is the internal pricing workbook that drives the Base Bid shown on
03 Bid Proposal §1. It lives in {Job}/Internal/BTQ/.

Workbook structure (8 sheets):
  1. Summary          ← Base Bid, Materials + Labor + Services buckets, SOV
  2. Materials        ← qty × Sale Price from Master Catalog Sheet 2
  3. Equipment        ← qty × Sale Price from Master Catalog Sheet 1
  4. Labor            ← hrs from Master Catalog Sheet 3 × buffer × $100/hr billed
  5. Services         ← labor + subs
  6. Alternates       ← ADDs (MS, accelerated, bonding) + DEDUCTs (after-hours credit)
  7. Unit Prices      ← 10 per-unit rates for change orders
  8. Milestones       ← 10/30/30/20/10 payment schedule

Governance compliance:
  - Cat6A plenum default on all data drops (no Cat6 substitute line items)
  - After-hours = DEDUCT (labor-surplus credit), never ADD premium
  - Billed labor $100/hr flat (foreman/laborer split never shown)
  - Internal cost rates: $65 foreman, $40 laborer, $52.50 blended (Summary sheet
    only; visible via hidden column for audit)
  - Buffer: 5% ≤50 drops · 10% 50-500 · 15% >500 or complex
  - Brand: Forest Green #2E7D32 headers, Slate #455A64 borders, Calibri 10pt

Governance source: 1-Bids/Rules.md, memory ## Estimation & Labor Rules
Pricing source of truth: 4-Company/Pricing/Master Catalog.xlsx

Output: AI_DRIVE_OUTPUT/BTQ - {Job Name}.xlsx

Author: GCC LV Div27-28
"""

from __future__ import annotations

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUTPUT_DIR = Path("/home/user/AI_DRIVE_OUTPUT")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ==========================================================================
# CONFIG — edit these per job
# ==========================================================================

CONFIG = {
    "job_name": "REPLACE-ME",
    "gc_name": "REPLACE-ME",
    "site": "REPLACE-ME",
    "city_state": "REPLACE-ME",
    "bid_date": "2026-04-20",

    # Drop count category (drives buffer tier)
    "total_drops": 0,            # ≤50 = 5%, 50-500 = 10%, >500 = 15%
    "complex_access": False,     # Occupied · night · high ceilings · outdoor · phased
    "prevailing_wage": False,    # Public jobs only — MO Wage Order 32

    # Materials (qty, unit, description, sale_each)
    "materials": [
        # Example:
        # (170*2, "lf", "Cat6A plenum blue (data)", 1.28),
    ],

    # Equipment
    "equipment": [
        # (1, "ea", "APC Netshelter SX 42U", 2850.00),
    ],

    # Labor lines (qty, unit, service, hrs_each)
    "labor": [
        # (170, "drop", "Cat6A drop install + terminate + certify + label", 1.05),
    ],

    # Managed Services tier (for §4 alternate) — None / "Small" / "Team" / "Enterprise" / "Mega"
    "ms_tier": None,

    # Soft costs
    "bond_pct": 0.0,             # 0.010 to 0.015 if bond required
    "ocip_rider_pct": 0.008,     # 0.008 to 0.012
    "mob_demob_hours": 20,       # 8 kickoff + 4/wk + 8 closeout
    "truck_transfers": 2,        # 4 hrs each
    "project_ship_cost": 0,      # Per-job shipping allowance

    # Contingency — carried INSIDE material/equipment subtotal, never disclosed
    "contingency_pct": 0.05,
}

# ==========================================================================
# Constants (governance source: memory ## Estimation & Labor Rules)
# ==========================================================================

LABOR_BILLED_RATE = 100.00          # $100/hr billed to client — never broken out
LABOR_COST_FOREMAN = 65.00           # Internal only
LABOR_COST_LABORER = 40.00           # Internal only
LABOR_COST_BLENDED = 52.50           # 1F+1L crew — internal default
LABOR_COST_PW = 63.26                # MO Wage Order 32 Jackson County Comm Tech

# Brand colors (hex without #, openpyxl style)
FOREST_GREEN = "2E7D32"
SLATE = "455A64"
WARM_GOLD = "D4AF37"
CREAM = "FDFBF4"
WHITE = "FFFFFF"

BODY_FONT = "Calibri"

# ==========================================================================
# Styling helpers
# ==========================================================================

thin_side = Side(border_style="thin", color=SLATE)
THIN_BORDER = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _header_style(cell):
    cell.font = Font(name=BODY_FONT, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=FOREST_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _body_style(cell, bold=False, italic=False, color=SLATE, align="left"):
    cell.font = Font(name=BODY_FONT, size=10, bold=bold, italic=italic, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _total_style(cell, color=FOREST_GREEN):
    cell.font = Font(name=BODY_FONT, size=10, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=CREAM)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = THIN_BORDER


def _money(cell, color=SLATE, bold=False):
    cell.font = Font(name=BODY_FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell.border = THIN_BORDER


# ==========================================================================
# Buffer logic
# ==========================================================================

def buffer_pct(total_drops: int, complex_access: bool) -> float:
    if complex_access or total_drops > 500:
        return 0.15
    if total_drops >= 50:
        return 0.10
    return 0.05


# ==========================================================================
# Sheet builders
# ==========================================================================

def build_materials_sheet(wb, cfg):
    ws = wb.create_sheet("Materials")
    headers = ["Qty", "Unit", "Description", "Sale $ Each", "Extended $"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        _header_style(c)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A2"

    row = 2
    for qty, unit, desc, sale in cfg["materials"]:
        ws.cell(row=row, column=1, value=qty); _body_style(ws.cell(row=row, column=1), align="center")
        ws.cell(row=row, column=2, value=unit); _body_style(ws.cell(row=row, column=2), align="center")
        ws.cell(row=row, column=3, value=desc); _body_style(ws.cell(row=row, column=3))
        c4 = ws.cell(row=row, column=4, value=sale); _money(c4)
        c5 = ws.cell(row=row, column=5, value=f"=A{row}*D{row}"); _money(c5)
        row += 1

    # Subtotal row
    last = row - 1
    ws.cell(row=row, column=3, value="Subtotal Materials")
    _total_style(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
    tot = ws.cell(row=row, column=5, value=f"=SUM(E2:E{last})" if last >= 2 else 0)
    _total_style(tot)
    tot.number_format = '"$"#,##0.00'

    return f"Materials!E{row}"


def build_equipment_sheet(wb, cfg):
    ws = wb.create_sheet("Equipment")
    headers = ["Qty", "Unit", "Description", "Sale $ Each", "Extended $"]
    for i, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=i, value=h))
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A2"

    row = 2
    for qty, unit, desc, sale in cfg["equipment"]:
        _body_style(ws.cell(row=row, column=1, value=qty), align="center")
        _body_style(ws.cell(row=row, column=2, value=unit), align="center")
        _body_style(ws.cell(row=row, column=3, value=desc))
        _money(ws.cell(row=row, column=4, value=sale))
        _money(ws.cell(row=row, column=5, value=f"=A{row}*D{row}"))
        row += 1

    last = row - 1
    ws.cell(row=row, column=3, value="Subtotal Equipment")
    _total_style(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
    tot = ws.cell(row=row, column=5, value=f"=SUM(E2:E{last})" if last >= 2 else 0)
    _total_style(tot)
    tot.number_format = '"$"#,##0.00'
    return f"Equipment!E{row}"


def build_labor_sheet(wb, cfg):
    ws = wb.create_sheet("Labor")
    headers = ["Qty", "Unit", "Service", "Hrs Each", "Total Hrs",
               "Billed $/hr", "Billed $", "Cost $/hr (int)", "Cost $ (int)"]
    for i, h in enumerate(headers, 1):
        _header_style(ws.cell(row=1, column=i, value=h))
    widths = [8, 8, 40, 10, 10, 12, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    cost_rate = LABOR_COST_PW if cfg.get("prevailing_wage") else LABOR_COST_BLENDED

    row = 2
    for qty, unit, desc, hrs_each in cfg["labor"]:
        _body_style(ws.cell(row=row, column=1, value=qty), align="center")
        _body_style(ws.cell(row=row, column=2, value=unit), align="center")
        _body_style(ws.cell(row=row, column=3, value=desc))
        _body_style(ws.cell(row=row, column=4, value=hrs_each), align="right")
        ws.cell(row=row, column=4).number_format = "0.00"
        _body_style(ws.cell(row=row, column=5, value=f"=A{row}*D{row}"), align="right")
        ws.cell(row=row, column=5).number_format = "0.0"
        _money(ws.cell(row=row, column=6, value=LABOR_BILLED_RATE))
        _money(ws.cell(row=row, column=7, value=f"=E{row}*F{row}"))
        _money(ws.cell(row=row, column=8, value=cost_rate))
        _money(ws.cell(row=row, column=9, value=f"=E{row}*H{row}"))
        row += 1

    last = row - 1
    # Subtotal
    ws.cell(row=row, column=3, value=f"Subtotal (catalog hours, no buffer)")
    _total_style(ws.cell(row=row, column=3))
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
    _total_style(ws.cell(row=row, column=5, value=f"=SUM(E2:E{last})" if last >= 2 else 0))
    ws.cell(row=row, column=5).number_format = "0.0"
    _total_style(ws.cell(row=row, column=7, value=f"=SUM(G2:G{last})" if last >= 2 else 0))
    ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
    _total_style(ws.cell(row=row, column=9, value=f"=SUM(I2:I{last})" if last >= 2 else 0))
    ws.cell(row=row, column=9).number_format = '"$"#,##0.00'

    # Buffer row
    buf_row = row + 1
    buf = buffer_pct(cfg["total_drops"], cfg["complex_access"])
    ws.cell(row=buf_row, column=3, value=f"Buffer ({int(buf*100)}% per Conservative Labor Rule)")
    _body_style(ws.cell(row=buf_row, column=3), italic=True, align="right")
    _body_style(ws.cell(row=buf_row, column=5, value=f"=E{row}*{buf}"), align="right")
    ws.cell(row=buf_row, column=5).number_format = "0.0"
    _money(ws.cell(row=buf_row, column=7, value=f"=G{row}*{buf}"))
    _money(ws.cell(row=buf_row, column=9, value=f"=I{row}*{buf}"))

    # Mob/demob row
    mob_row = buf_row + 1
    mob_hrs = cfg["mob_demob_hours"] + cfg["truck_transfers"] * 4
    ws.cell(row=mob_row, column=3, value=f"Mobilization / demobilization ({mob_hrs} hrs)")
    _body_style(ws.cell(row=mob_row, column=3), align="right")
    _body_style(ws.cell(row=mob_row, column=5, value=mob_hrs), align="right")
    ws.cell(row=mob_row, column=5).number_format = "0.0"
    _money(ws.cell(row=mob_row, column=7, value=mob_hrs * LABOR_BILLED_RATE))
    _money(ws.cell(row=mob_row, column=9, value=mob_hrs * cost_rate))

    # Grand total row
    gt_row = mob_row + 1
    ws.cell(row=gt_row, column=3, value="TOTAL LABOR (billed)")
    _total_style(ws.cell(row=gt_row, column=3))
    ws.cell(row=gt_row, column=3).alignment = Alignment(horizontal="right", vertical="center")
    _total_style(ws.cell(row=gt_row, column=5, value=f"=E{row}+E{buf_row}+E{mob_row}"))
    ws.cell(row=gt_row, column=5).number_format = "0.0"
    _total_style(ws.cell(row=gt_row, column=7, value=f"=G{row}+G{buf_row}+G{mob_row}"))
    ws.cell(row=gt_row, column=7).number_format = '"$"#,##0.00'
    _total_style(ws.cell(row=gt_row, column=9, value=f"=I{row}+I{buf_row}+I{mob_row}"))
    ws.cell(row=gt_row, column=9).number_format = '"$"#,##0.00'

    return f"Labor!G{gt_row}", f"Labor!I{gt_row}"


def build_services_sheet(wb):
    """Services sheet — typically empty unless bid includes specific services (rare)."""
    ws = wb.create_sheet("Services")
    _header_style(ws.cell(row=1, column=1, value="Service"))
    _header_style(ws.cell(row=1, column=2, value="Unit"))
    _header_style(ws.cell(row=1, column=3, value="Qty"))
    _header_style(ws.cell(row=1, column=4, value="Sale $"))
    _header_style(ws.cell(row=1, column=5, value="Extended $"))
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Subtotal row at row 10 to leave room for adds
    _total_style(ws.cell(row=10, column=1, value="Subtotal Services"))
    ws.cell(row=10, column=1).alignment = Alignment(horizontal="right", vertical="center")
    _total_style(ws.cell(row=10, column=5, value="=SUM(E2:E9)"))
    ws.cell(row=10, column=5).number_format = '"$"#,##0.00'
    return "Services!E10"


def build_alternates_sheet(wb, cfg):
    ws = wb.create_sheet("Alternates")
    for i, h in enumerate(["Ref", "Description", "Type", "Amount"], 1):
        _header_style(ws.cell(row=1, column=i, value=h))
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    rows = [
        ("A1", f"UniFi Managed Services — {cfg.get('ms_tier') or 'Team'} tier (hardware + setup + first year recurring)", "ADD", ""),
        ("A2", "Accelerated schedule — increased crew + overlapping phases", "ADD", ""),
        ("A3", "Performance & payment bonding (100/100)", "ADD", ""),
        ("D1", "Weekend + after-hours scheduling — GCC labor-surplus credit (DEDUCT)", "DEDUCT", ""),
    ]
    for i, (ref, desc, kind, amt) in enumerate(rows, 2):
        _body_style(ws.cell(row=i, column=1, value=ref), align="center")
        _body_style(ws.cell(row=i, column=2, value=desc))
        color = FOREST_GREEN if kind == "ADD" else WARM_GOLD
        _body_style(ws.cell(row=i, column=3, value=kind), bold=True, color=color, align="center")
        if amt == "":
            _body_style(ws.cell(row=i, column=4, value=""), align="right")
        else:
            _money(ws.cell(row=i, column=4, value=amt))


def build_unit_prices_sheet(wb):
    ws = wb.create_sheet("Unit Prices")
    for i, h in enumerate(["Ref", "Description", "Unit", "Rate $"], 1):
        _header_style(ws.cell(row=1, column=i, value=h))
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    rows = [
        ("U1", "Cat6A plenum drop — installed, terminated both ends, certified, labeled", "per drop", ""),
        ("U2", "Voice drop (Cat6A plenum) — installed, terminated, tested", "per drop", ""),
        ("U3", "Fiber strand termination (SC/LC fusion)", "per strand", ""),
        ("U4", "IP camera install (mount + aim + test)", "per device", ""),
        ("U5", "WAP install (mount + commission)", "per device", ""),
        ("U6", "ACS reader install (wall mount + wire + terminate)", "per device", ""),
        ("U7", "Firestop penetration beyond allowance", "per each", ""),
        ("U8", "Patch panel port (punchdown + label)", "per port", ""),
        ("U9", "Additional mobilization / remobilization", "per each", ""),
        ("U10", "Change-order labor (T&M)", "per hour", LABOR_BILLED_RATE),
    ]
    for i, (ref, desc, unit, rate) in enumerate(rows, 2):
        _body_style(ws.cell(row=i, column=1, value=ref), align="center")
        _body_style(ws.cell(row=i, column=2, value=desc))
        _body_style(ws.cell(row=i, column=3, value=unit), align="center")
        if rate == "":
            _body_style(ws.cell(row=i, column=4, value=""), align="right")
        else:
            _money(ws.cell(row=i, column=4, value=rate))


def build_milestones_sheet(wb):
    ws = wb.create_sheet("Milestones")
    for i, h in enumerate(["Milestone", "Description", "% of Base Bid"], 1):
        _header_style(ws.cell(row=1, column=i, value=h))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16

    rows = [
        ("M1", "Mobilization + material release", "10%"),
        ("M2", "Rough-in 50% complete (pathway + backboxes)", "30%"),
        ("M3", "Rough-in complete (all cabling pulled)", "30%"),
        ("M4", "Termination + testing complete", "20%"),
        ("M5", "Substantial Completion + closeout delivered", "10%"),
    ]
    for i, (ms, desc, pct) in enumerate(rows, 2):
        _body_style(ws.cell(row=i, column=1, value=ms), bold=True, align="center")
        _body_style(ws.cell(row=i, column=2, value=desc))
        _body_style(ws.cell(row=i, column=3, value=pct), bold=True, align="center")


def build_summary_sheet(wb, cfg, mat_cell, eq_cell, labor_billed_cell, labor_cost_cell, svc_cell):
    ws = wb.create_sheet("Summary", 0)  # Position as first sheet

    # Title block
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1, value=f"BTQ — {cfg['job_name']}")
    c.font = Font(name=BODY_FONT, size=16, bold=True, color=FOREST_GREEN)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1, value=f"{cfg['gc_name']}  ·  {cfg['city_state']}  ·  Bid {cfg['bid_date']}")
    c.font = Font(name=BODY_FONT, size=10, italic=True, color=SLATE)

    ws.merge_cells("A3:E3")
    c = ws.cell(row=3, column=1,
                value="GCC Proprietary & Confidential  ·  Internal pricing document  ·  Not for GC or Owner distribution")
    c.font = Font(name=BODY_FONT, size=9, italic=True, color="CC0000")

    # Pricing rollup header
    for i, h in enumerate(["Bucket", "Description", "", "", "Amount $"], 1):
        cell = ws.cell(row=5, column=i, value=h)
        _header_style(cell)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16

    buckets = [
        ("Materials", "Cable, jacks, faceplates, pathway, firestop, labels, misc hardware", f"={mat_cell}+={eq_cell}"),
        ("Labor",     "All labor + buffer + mob/demob (billed at $100/hr flat)", f"={labor_billed_cell}"),
        ("Services",  "Labor-based services + subs (if any)", f"={svc_cell}"),
    ]
    row = 6
    for bucket, desc, formula in buckets:
        _body_style(ws.cell(row=row, column=1, value=bucket), bold=True)
        _body_style(ws.cell(row=row, column=2, value=desc))
        c = ws.cell(row=row, column=5, value=formula)
        _money(c)
        row += 1

    # Base Bid row
    _total_style(ws.cell(row=row, column=1, value="BASE BID"))
    _total_style(ws.cell(row=row, column=2, value="Rounded to nearest $100 on Bid Proposal §1"))
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
    bb = ws.cell(row=row, column=5, value=f"=SUM(E6:E{row-1})")
    bb.font = Font(name=BODY_FONT, size=12, bold=True, color=FOREST_GREEN)
    bb.fill = PatternFill("solid", fgColor=CREAM)
    bb.alignment = Alignment(horizontal="right", vertical="center")
    bb.number_format = '"$"#,##0.00'
    bb.border = THIN_BORDER
    base_row = row

    # Internal margin analysis (hidden-ish — far right, italic, red)
    row += 3
    _header_style(ws.cell(row=row, column=1, value="INTERNAL MARGIN ANALYSIS — DO NOT SHARE"))
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="CC0000")
    row += 1
    _body_style(ws.cell(row=row, column=1, value="Labor Cost (internal)"))
    _money(ws.cell(row=row, column=5, value=f"={labor_cost_cell}"))
    row += 1
    _body_style(ws.cell(row=row, column=1, value="Gross Margin $ (Base Bid − COGS)"))
    _money(ws.cell(row=row, column=5, value=f"=E{base_row}-E{row-1}"))
    row += 1
    _body_style(ws.cell(row=row, column=1, value="Gross Margin %"), italic=True)
    gm = ws.cell(row=row, column=5, value=f"=(E{base_row}-E{row-2})/E{base_row}")
    gm.font = Font(name=BODY_FONT, size=10, bold=True, italic=True, color=SLATE)
    gm.number_format = "0.0%"
    gm.alignment = Alignment(horizontal="right", vertical="center")
    gm.border = THIN_BORDER

    # Conditional format — red if <28%, green if ≥35%
    ws.conditional_formatting.add(f"E{row}",
        CellIsRule(operator="lessThan", formula=["0.28"], fill=PatternFill(bgColor="FFCDD2")))
    ws.conditional_formatting.add(f"E{row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.35"], fill=PatternFill(bgColor="C8E6C9")))


# ==========================================================================
# Main
# ==========================================================================

def generate_btq(config=None):
    cfg = config or CONFIG
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    mat_ref = build_materials_sheet(wb, cfg)
    eq_ref = build_equipment_sheet(wb, cfg)
    lab_billed_ref, lab_cost_ref = build_labor_sheet(wb, cfg)
    svc_ref = build_services_sheet(wb)
    build_alternates_sheet(wb, cfg)
    build_unit_prices_sheet(wb)
    build_milestones_sheet(wb)

    # Summary last, placed first in tab order
    build_summary_sheet(wb, cfg, mat_ref, eq_ref, lab_billed_ref, lab_cost_ref, svc_ref)

    # Output
    outpath = OUTPUT_DIR / f"BTQ - {cfg['job_name']}.xlsx"
    wb.save(outpath)
    print(f"[OK] BTQ generated: {outpath.name}")
    print(f"     Sheets: Summary · Materials · Equipment · Labor · Services · Alternates · Unit Prices · Milestones")
    print(f"     Buffer: {int(buffer_pct(cfg['total_drops'], cfg['complex_access'])*100)}%")
    print(f"     Labor cost rate: ${LABOR_COST_PW if cfg.get('prevailing_wage') else LABOR_COST_BLENDED}/hr")
    return outpath


if __name__ == "__main__":
    generate_btq()
