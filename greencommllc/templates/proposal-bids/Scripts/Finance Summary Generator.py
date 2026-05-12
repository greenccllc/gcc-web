"""
Finance Summary Generator — GCC LV Div27-28 — INTERNAL ONLY

Generates a per-bid P&L scenario workbook. Models base / efficient / stressed
scenarios and calculates cost breakdown, labor cost vs billed, bonuses,
and margin. Feeds the Finance Summary narrative DOCX.

This output is internal only. Never ships to GC, Owner, or sub.

Governance source: 1-Bids/Rules.md, memory ## Estimation & Labor Rules.
Pricing source of truth: 4-Company/Pricing/Master Catalog.xlsx
Paired narrative: 1-Bids/Templates/Internal-Only/Finance Summary.docx

Workbook structure:
  1. Cover         — job identity, internal banner
  2. Cost Breakdown — Materials / Equipment / Labor with cost, sale, margin
  3. Scenarios     — base / efficient (+10%) / stressed (-15%) / prevailing-wage
  4. Margin Trend  — margin table vs target band (28-35%)
  5. Change Orders — running CO log + CO margin

Output: AI_DRIVE_OUTPUT/Finance Summary - {Job Name}.xlsx

Author: GCC LV Div27-28
"""

from __future__ import annotations

import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

OUTPUT_DIR = Path("/home/user/AI_DRIVE_OUTPUT")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ==========================================================================
# CONFIG
# ==========================================================================

CONFIG = {
    "job_name": "REPLACE-ME",
    "gc_name": "REPLACE-ME",
    "site": "REPLACE-ME",
    "city_state": "REPLACE-ME",
    "bid_date": "2026-04-20",
    "base_bid": 0.00,              # From BTQ Summary sheet
    "materials_cost": 0.00,
    "materials_sale": 0.00,
    "equipment_cost": 0.00,
    "equipment_sale": 0.00,
    "labor_hours": 0.0,            # Buffered hours
    "prevailing_wage": False,
    "managed_services_in_base": False,  # True for whole-building jobs
    "bond_pct": 0.0,
}

# Constants
BILLED_RATE = 100.00
COST_BLENDED = 52.50
COST_PW = 63.26
MARGIN_TARGET_LOW = 0.28
MARGIN_TARGET_HIGH = 0.35
MARGIN_FLOOR = 0.28

FOREST_GREEN = "2E7D32"
SLATE = "455A64"
WARM_GOLD = "D4AF37"
CREAM = "FDFBF4"
WHITE = "FFFFFF"
RED = "CC0000"
BODY_FONT = "Calibri"

thin_side = Side(border_style="thin", color=SLATE)
THIN = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)


def _hdr(cell):
    cell.font = Font(name=BODY_FONT, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=FOREST_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = THIN


def _body(cell, bold=False, color=SLATE, align="left", italic=False):
    cell.font = Font(name=BODY_FONT, size=10, bold=bold, italic=italic, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = THIN


def _money(cell, bold=False, color=SLATE):
    cell.font = Font(name=BODY_FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '"$"#,##0.00'
    cell.border = THIN


def _pct(cell, bold=False, color=SLATE):
    cell.font = Font(name=BODY_FONT, size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = "0.0%"
    cell.border = THIN


# ==========================================================================
# Sheets
# ==========================================================================

def build_cover(wb, cfg):
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 58

    # Internal banner
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1,
                value="[INTERNAL ONLY]  GCC Proprietary & Confidential  ·  Never share with GC, Owner, or sub")
    c.font = Font(name=BODY_FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=RED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:B3")
    c = ws.cell(row=3, column=1, value=f"Finance Summary — {cfg['job_name']}")
    c.font = Font(name=BODY_FONT, size=16, bold=True, color=FOREST_GREEN)

    rows = [
        ("GC", cfg["gc_name"]),
        ("Site", cfg["site"]),
        ("Location", cfg["city_state"]),
        ("Bid Date", cfg["bid_date"]),
        ("Base Bid", cfg["base_bid"]),
        ("Prevailing Wage", "YES" if cfg["prevailing_wage"] else "NO"),
        ("Managed Services in Base", "YES" if cfg["managed_services_in_base"] else "NO"),
    ]
    r = 5
    for k, v in rows:
        _body(ws.cell(row=r, column=1, value=k), bold=True, color=FOREST_GREEN)
        if k == "Base Bid":
            _money(ws.cell(row=r, column=2, value=v))
        else:
            _body(ws.cell(row=r, column=2, value=v))
        r += 1


def build_cost_breakdown(wb, cfg):
    ws = wb.create_sheet("Cost Breakdown")
    cost_rate = COST_PW if cfg["prevailing_wage"] else COST_BLENDED

    for i, h in enumerate(["Category", "Cost $", "Sale $", "Margin $", "Margin %"], 1):
        _hdr(ws.cell(row=1, column=i, value=h))
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    mat_cost = cfg["materials_cost"]
    mat_sale = cfg["materials_sale"]
    eq_cost = cfg["equipment_cost"]
    eq_sale = cfg["equipment_sale"]
    lab_hrs = cfg["labor_hours"]
    lab_cost = lab_hrs * cost_rate
    lab_sale = lab_hrs * BILLED_RATE

    rows = [
        ("Materials (Master Catalog Sheet 2)", mat_cost, mat_sale),
        ("Equipment (Master Catalog Sheet 1)", eq_cost, eq_sale),
        (f"Labor ({lab_hrs:.1f} hrs @ ${cost_rate} cost / ${BILLED_RATE} billed)", lab_cost, lab_sale),
    ]

    r = 2
    for label, cost, sale in rows:
        _body(ws.cell(row=r, column=1, value=label))
        _money(ws.cell(row=r, column=2, value=cost))
        _money(ws.cell(row=r, column=3, value=sale))
        _money(ws.cell(row=r, column=4, value=f"=C{r}-B{r}"))
        _pct(ws.cell(row=r, column=5, value=f"=IFERROR((C{r}-B{r})/C{r},0)"))
        r += 1

    # Total
    _body(ws.cell(row=r, column=1, value="TOTAL PROJECT"), bold=True, color=FOREST_GREEN)
    _money(ws.cell(row=r, column=2, value=f"=SUM(B2:B{r-1})"), bold=True, color=FOREST_GREEN)
    _money(ws.cell(row=r, column=3, value=f"=SUM(C2:C{r-1})"), bold=True, color=FOREST_GREEN)
    _money(ws.cell(row=r, column=4, value=f"=SUM(D2:D{r-1})"), bold=True, color=FOREST_GREEN)
    _pct(ws.cell(row=r, column=5, value=f"=IFERROR((C{r}-B{r})/C{r},0)"), bold=True, color=FOREST_GREEN)

    # Color the margin cell per target band
    ws.conditional_formatting.add(f"E{r}",
        CellIsRule(operator="lessThan", formula=[str(MARGIN_FLOOR)], fill=PatternFill(bgColor="FFCDD2")))
    ws.conditional_formatting.add(f"E{r}",
        CellIsRule(operator="greaterThanOrEqual", formula=[str(MARGIN_TARGET_HIGH)], fill=PatternFill(bgColor="C8E6C9")))

    return r  # totals row number


def build_scenarios(wb, cfg, totals_row):
    ws = wb.create_sheet("Scenarios")
    for i, h in enumerate(["Scenario", "Revenue $", "Cost $", "Margin $", "Margin %", "Notes"], 1):
        _hdr(ws.cell(row=1, column=i, value=h))
    widths = [24, 14, 14, 14, 12, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Reference Cost Breakdown totals
    rev = f"='Cost Breakdown'!C{totals_row}"
    cost = f"='Cost Breakdown'!B{totals_row}"

    scenarios = [
        ("Base Case (BTQ as bid)", rev, cost,
         "Scope and labor as priced; no efficiency wins, no overrun"),
        ("Efficient (+10% productivity)", rev, f"({cost})*0.90",
         "Crew hits stretch labor targets — 10% under planned hours"),
        ("Stressed (-15% productivity)", rev, f"({cost})*1.15",
         "Access delays, RFI churn, minor scope creep absorbed"),
        ("PW applied (MO Wage Order 32)", rev,
         f"={cost}-'Cost Breakdown'!B4+('Cost Breakdown'!B4/{COST_BLENDED})*{COST_PW}",
         "Applies $48.33 PW to labor hours — MO public jobs only"),
        ("CO +20% at Base margin", f"({rev})*1.20", f"({cost})*1.20",
         "Assumes 20% CO at the same margin rate — healthy upside"),
    ]

    r = 2
    for name, rev_f, cost_f, note in scenarios:
        _body(ws.cell(row=r, column=1, value=name), bold=True, color=FOREST_GREEN)
        _money(ws.cell(row=r, column=2, value=f"={rev_f}" if rev_f.startswith("=") or not rev_f.startswith("(") else f"={rev_f}"))
        _money(ws.cell(row=r, column=3, value=f"={cost_f}" if cost_f.startswith("=") or not cost_f.startswith("(") else f"={cost_f}"))
        _money(ws.cell(row=r, column=4, value=f"=B{r}-C{r}"))
        mgn = ws.cell(row=r, column=5, value=f"=IFERROR((B{r}-C{r})/B{r},0)")
        _pct(mgn)
        _body(ws.cell(row=r, column=6, value=note), italic=True)
        ws.conditional_formatting.add(f"E{r}",
            CellIsRule(operator="lessThan", formula=[str(MARGIN_FLOOR)], fill=PatternFill(bgColor="FFCDD2")))
        ws.conditional_formatting.add(f"E{r}",
            CellIsRule(operator="greaterThanOrEqual", formula=[str(MARGIN_TARGET_HIGH)], fill=PatternFill(bgColor="C8E6C9")))
        r += 1


def build_margin_trend(wb):
    ws = wb.create_sheet("Margin Trend")
    _hdr(ws.cell(row=1, column=1, value="Target Band"))
    _hdr(ws.cell(row=1, column=2, value="Range"))
    _hdr(ws.cell(row=1, column=3, value="Action"))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 58

    rows = [
        ("HARD FLOOR", "< 28%", "Reconsider or WALK. Margin below floor — likely pricing error or scope underbid."),
        ("TARGET BAND", "28% – 35%", "Healthy. Proceed to bid."),
        ("PREMIUM", "> 35%", "Above target — verify this is premium pricing, not a pricing error."),
        ("ALARM ZONE", "> 50%", "Reprice. Likely duplicated line items or missing cost categories."),
    ]
    r = 2
    for band, rng, action in rows:
        color = RED if band == "HARD FLOOR" else (FOREST_GREEN if band == "TARGET BAND" else SLATE)
        _body(ws.cell(row=r, column=1, value=band), bold=True, color=color)
        _body(ws.cell(row=r, column=2, value=rng), bold=True, color=color, align="center")
        _body(ws.cell(row=r, column=3, value=action))
        r += 1


def build_co_log(wb):
    ws = wb.create_sheet("Change Orders")
    for i, h in enumerate(["CO #", "Date", "Description", "Revenue $", "Cost $", "Margin $", "Margin %", "Status"], 1):
        _hdr(ws.cell(row=1, column=i, value=h))
    widths = [8, 12, 40, 14, 14, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Empty rows for tracking
    for r in range(2, 22):
        for col in range(1, 9):
            ws.cell(row=r, column=col, value="").border = THIN


# ==========================================================================
# Main
# ==========================================================================

def generate_finance_summary(config=None):
    cfg = config or CONFIG
    wb = Workbook()

    build_cover(wb, cfg)
    totals_row = build_cost_breakdown(wb, cfg)
    build_scenarios(wb, cfg, totals_row)
    build_margin_trend(wb)
    build_co_log(wb)

    outpath = OUTPUT_DIR / f"Finance Summary - {cfg['job_name']}.xlsx"
    wb.save(outpath)
    print(f"[OK] Finance Summary generated: {outpath.name}")
    print(f"     Internal only — never shared with GC or Owner")
    return outpath


if __name__ == "__main__":
    generate_finance_summary()
